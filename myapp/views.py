from django.shortcuts import render, HttpResponse, redirect, reverse
from django.http import HttpResponseRedirect, JsonResponse
from .models import dependent, beneficiary, beneficiary_house, beneficiary_income_expense, Dependent_income, Beneficiary_attachment, Supporter_beneficiary_sponsorship, CustomUser, Beneficiary_request, Supporter, Supporter_request, Supporter_request_attachment, Support_operation, Support_operation_attachment, Field_visit, Field_visit_attachment, Authentication_OTP
# from .forms import CustomUserCreationForm
# from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password
from django.contrib.auth import login, logout, authenticate
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import logging
import os
import re
import json
import time
import random
import requests
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import *
import decimal
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.mail import send_mail, EmailMessage, BadHeaderError
from cfc_app import settings
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from .tokens import generate_token
from django.utils.timezone import now
from django.contrib.auth.tokens import default_token_generator
from django.shortcuts import get_object_or_404
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.models import User, Group
from django.db.models.query_utils import Q
from django.db.models import Sum, Avg, Count, F
from .decorators import group_required
from dotenv import load_dotenv

load_dotenv()

# Set up logging
logger = logging.getLogger(__name__)

# Constants =======================================

IPP_DASHBOARD_REQUESTS = 10  # IPP stands for Item Per Page
IPP_SUPPORTER_FORM = 8
IPP_DASHBOARD_REPORTS = 10
# Maximum messages allowed per minute
SMS_LIMIT_PER_MINUTE = 5

duration_factors = {
    "": 0,
    "شهر واحد": 1,
    "3 أشهر": 3,
    "6 أشهر": 6,
    "سنة كاملة": 12,
    # Add more duration types and their factors as needed
}

# Utility functions =======================================


def file_extension(value):
    _, extension = os.path.splitext(value)
    return extension.lower()


def convert_to_date(date_str):
    if not date_str:
        return None  # or handle it as needed in your context

    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        print(f"Invalid date format for {date_str}")
        return None  # or raise an exception, depending on your requirement


def is_valid_queryparam(param, type):
    if type == 1:
        return param != '' and param is not None
    elif type == 2:
        return param != "اختار..." and param is not None


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


def generate_otp():
    """Generate a 6-digit OTP code."""
    return random.randint(100000, 999999)


def normalize_phone_number(phone):
    """
    Normalize the phone number entered by the user into the format used in the database (e.g., 541546323).
    
    Accepted formats:
    - +9665XXXXXXXX
    - 9665XXXXXXXX
    - 5XXXXXXXX

    Returns:
    - The normalized phone number (e.g., 541546323) if valid.
    - None if the phone number is invalid.
    """
    # Remove any non-numeric characters (if necessary)
    phone = ''.join(filter(str.isdigit, phone))

    # Check and normalize based on known patterns
    if phone.startswith('+9665') and len(phone) == 13:  # +9665XXXXXXXX
        return phone[4:]  # Remove '+966'
    elif phone.startswith('9665') and len(phone) == 12:  # 9665XXXXXXXX
        return phone[3:]  # Remove '966'
    elif phone.startswith('5') and len(phone) == 9:  # 5XXXXXXXX
        return phone  # Already in the correct format

    # If none of the patterns match, return None (invalid phone number)
    return None


def send_otp_via_sms(phone_number, otp, user):
    api_url = os.environ.get('SMS_PROVIDER_API_BASE_URL') + "/api/v1/sendsms"
    api_key = os.environ.get('SMS_PROVIDER_API_KEY')
    username = os.environ.get('SMS_PROVIDER_USERNAME')
    sender = os.environ.get('SMS_PROVIDER_SENDER')
    message = f"رمز التحقق هو: {otp}"

    # Ensure the phone number starts with "966"
    if not phone_number.startswith("966"):
        phone_number = "966" + phone_number

    # Get the current time and calculate the start of the one-minute window
    current_time = now()
    one_minute_ago = current_time - timedelta(minutes=1)

    # Check how many OTPs have been sent by this user in the last minute
    recent_otps = Authentication_OTP.objects.filter(
        created_by=user,
        created_at__gte=one_minute_ago
    ).count()

    print("SMS is processing....")

    if recent_otps >= SMS_LIMIT_PER_MINUTE:
        print(f"SMS limit reached for user {user.id}. Try again later.")
        return False

    params = {
        "api_key": api_key,
        "username": username,
        "numbers": phone_number,
        "response_type": "text",
        "message": message,
        "sender": sender,
    }

    try:
        response = requests.get(api_url, params=params)
        print("SMS is about to be sent...")
        if response.status_code == 200:
            print(f"OTP sent successfully to {phone_number}")
            return True
        else:
            print(f"Failed to send OTP. Response: {response.text}")
            return False
    except requests.RequestException as e:
        print(f"Error occurred while sending OTP: {e}")
        return False

# View Handlers ==============================================


def forbidden(request):
    return render(request, 'errors/403.html')


def home(request):
    return render(request, "home.html")


def home_redirect(request):
    return redirect("/home")


def confirmBeneficiaryRequestView(request):
    return render(request, "main/confirmBeneficiaryReq.html")


def send_otp_email(user, otp_code):
    """
    Sends an OTP email to the user.
    """
    email_subject = "رمز التحقق الخاص بك - جمعية أصدقاء المجتمع"
    
    # Render email body using an HTML template
    message = render_to_string('auth/otp_template.html', {
        'name': user.first_name,
        'otp_code': otp_code,
        'expiry_minutes': 5,  # Adjust this value to match your expiry time
    })

    email = EmailMessage(
        email_subject,
        message,
        settings.EMAIL_HOST_USER,
        [user.email],
    )
    email.content_subtype = 'html'  # Specify that the email content is HTML
    email.fail_silently = False  # Raise errors if email sending fails
    email.send()


def sign_up(request):
    if request.method == 'POST':

        # Check if the user is already authenticated
        if request.user.is_authenticated:
            return redirect('home')

        # Retrieve request data
        first_name = request.POST['first_name']
        second_name = request.POST['second_name']
        last_name = request.POST['last_name']
        date_of_birth = request.POST['date_of_birth']
        gender = request.POST['gender']
        national_id = request.POST['national_id']
        national_id_exp_date = request.POST['national_id_exp_date']
        nationality = request.POST['nationality']
        username = request.POST['username']
        email = request.POST['email']
        phonenumber = request.POST.get('phonenumber', None)
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        # Ensure no fields are empty
        if not all([first_name, second_name, last_name, date_of_birth, gender, national_id, national_id_exp_date, nationality, username, email, password1, password2]):
            messages.error(request, "جميع الحقول مطلوبة.")
            return redirect('sign-up')

        username_exists = CustomUser.objects.filter(username=username).exists()

        if username_exists:
            # Pass request data back
            request.session['first_name'] = first_name
            request.session['second_name'] = second_name
            request.session['last_name'] = last_name
            request.session['date_of_birth'] = date_of_birth
            request.session['gender'] = gender
            request.session['national_id'] = national_id
            request.session['national_id_exp_date'] = national_id_exp_date
            request.session['nationality'] = nationality
            request.session['username'] = username
            request.session['email'] = email
            request.session['phonenumber'] = phonenumber

            # Alert an error message
            messages.error(request, "أسم المستخدم موجود سابقًا.")

            return redirect('sign-up')

        email_exists = CustomUser.objects.filter(email__iexact=email).exists()

        if email_exists:
            # Pass request data back
            request.session['first_name'] = first_name
            request.session['second_name'] = second_name
            request.session['last_name'] = last_name
            request.session['date_of_birth'] = date_of_birth
            request.session['gender'] = gender
            request.session['national_id'] = national_id
            request.session['nationality'] = nationality
            request.session['username'] = username
            request.session['email'] = email
            request.session['phonenumber'] = phonenumber

            # Alert an error message
            messages.error(request, "البريد الالكتروني موجود سابقًا.")

            return redirect('sign-up')
        
        # Ensure passwords match
        if password1 != password2:
            messages.error(request, "كلمات المرور غير متطابقة.")
            return redirect('sign-up')
        
        # Validate date of birth (check format and ensure it's a valid date)
        try:
            dob = datetime.strptime(date_of_birth, "%Y-%m-%d")
        except ValueError:
            messages.error(request, "صيغة تاريخ الميلاد غير صحيحة.")
            return redirect('sign-up')

        # Make a new user object
        new_user = CustomUser(
            username=username,
            email=email,
            first_name=first_name,
            second_name=second_name,
            last_name=last_name,
            gender=gender,
            national_id=national_id,
            national_id_exp_date=national_id_exp_date,
            nationality=nationality,
            phonenumber=phonenumber,
            date_of_birth=date_of_birth,
            # Make the user not active until it is confirmed by the link sent to the email
            is_active=False,
        )
        # Set the password to be hashed for the new user
        new_user.set_password(password1)

        # Save new user information
        new_user.save()

        # Assign default role to the new user
        default_role = Group.objects.filter(
            name__iexact='beneficiary').first()

        # Give the new user the default role
        new_user.groups.add(default_role)

        messages.success(
             request, "تم إنشاء حسابك بنجاح! رجاء راجع البريد الالكتروني الخاص بك لتأكيد البريد الالكتروني وتفعيل حسابك.")

        # Pass user identifier to the otp verification page to let the user login using it
        request.session['username'] = username

        # Generate and save OTP
        otp_code = generate_otp()
        expiry_time = now() + timedelta(minutes=5)  # OTP expires in 5 minutes
        ip_address = request.META.get('REMOTE_ADDR', None)
        user_agent = request.META.get('HTTP_USER_AGENT', '')

        otp = Authentication_OTP(
            otp_code=otp_code,
            expiry_time=expiry_time,
            created_by=new_user,
            purpose="Sign-up phone number verification",
            ip_address=ip_address,
            user_agent=user_agent
        )
        otp.save()

        # Send OTP via email
        send_otp_email(new_user, otp.otp_code)

        return redirect('otp_sign_up_view')

    else:

        # Check if the user is already authenticated
        if request.user.is_authenticated:
            return redirect('home')

    return render(request, "registration/sign_up.html")


def otp_sign_up_view(request):
    if request.method == "POST":
        username = request.session.get('username', '')
        otp_code = request.POST.get('otp', '')

        if 'resend_otp' in request.POST:  # Handle resend OTP action
            try:
                user = CustomUser.objects.get(username=username)

                # Check if there's a recent OTP and update its expiry time
                try:
                    recent_otp = Authentication_OTP.objects.filter(
                        created_by=user,
                        purpose="Sign-up phone number verification",
                    ).latest('created_at')

                    if recent_otp.created_at + timedelta(minutes=1) > now():
                        messages.error(request, "يمكنك إعادة إرسال رمز التحقق بعد دقيقة واحدة.")
                        return redirect('otp_sign_up_view')

                    # Mark the recent OTP as expired
                    recent_otp.expiry_time = now()
                    recent_otp.save()
                except Authentication_OTP.DoesNotExist:
                    # No previous OTP exists, continue to generate a new one
                    pass

                # Generate and save new OTP
                new_otp_code = generate_otp()
                expiry_time = now() + timedelta(minutes=5)
                ip_address = request.META.get('REMOTE_ADDR', None)
                user_agent = request.META.get('HTTP_USER_AGENT', '')

                new_otp = Authentication_OTP(
                    otp_code=new_otp_code,
                    expiry_time=expiry_time,
                    created_by=user,
                    purpose="Sign-up phone number verification",
                    ip_address=ip_address,
                    user_agent=user_agent
                )
                new_otp.save()

                # Send OTP via email
                send_otp_email(user, new_otp_code)

                messages.success(request, "تم إرسال رمز تحقق جديد إلى رقم الجوال الخاص بك.")
            except CustomUser.DoesNotExist:
                messages.error(request, "حدث خطأ أثناء إرسال رمز التحقق.")
            return redirect('otp_sign_up_view')

        # Verify OTP
        try:
            user = CustomUser.objects.get(username=username)
            otp = Authentication_OTP.objects.filter(
                created_by=user,
                purpose="Sign-up phone number verification",
                is_used=False
            ).latest('created_at')

            # Check OTP validity
            if otp.otp_code == otp_code and otp.expiry_time > now():
                otp.is_used = True
                otp.used_at = now()
                otp.save()

                user.is_active = True
                user.save()
                messages.success(request, "تم تفعيل الحساب بنجاح! يمكنك الأن تسجيل الدخول باستخدام حسابك.")
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, "رمز التحقق غير صالح أو منتهي الصلاحية، الرجاء المحاولة مره أخرى.")
        except (CustomUser.DoesNotExist, Authentication_OTP.DoesNotExist):
            messages.error(request, "حدث خطأ أثناء التحقق.")
            return redirect('sign-up')

    elif request.method == "GET":
        username = request.session.get('username', '')
        return render(request, 'auth/otp_sign_up_view.html')

    return render(request, 'auth/otp_sign_up_view.html')


# def activate(request, uidb64, token):

#     try:
#         uid = force_str(urlsafe_base64_decode(uidb64))
#         myuser = CustomUser.objects.get(pk=uid)

#     except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist):
#         print("User doesn't exist")
#         myuser = None

#     if myuser is not None and generate_token.check_token(myuser, token):
#         myuser.is_active = True
#         myuser.save()
#         login(request, myuser)
#         return redirect('home')
#     else:
#         return render(request, 'auth/activation_failed.html')


def resend_activation_email_view(request):

    if request.method == 'POST':

        # Check if the user is already authenticated
        if request.user.is_authenticated:
            return redirect('home')

        email = request.POST.get("email")

        try:
            myuser = CustomUser.objects.get(email__iexact=email)

            if myuser is not None and myuser.is_active is False:
                # Generate a new OTP
                otp_code = generate_otp()
                expiry_time = now() + timedelta(minutes=5)

                # Get the user's IP address
                ip_address = request.META.get('REMOTE_ADDR', None)

                # Save the OTP to the database
                Authentication_OTP.objects.create(
                    otp_code=otp_code,
                    expiry_time=expiry_time,
                    created_by=myuser,
                    purpose="Account Activation",
                    ip_address=ip_address,
                )

                # Render the OTP email template
                email_subject = "رمز التحقق لتفعيل الحساب"
                email_message = render_to_string('auth/otp_template.html', {
                    'name': myuser.first_name,
                    'otp_code': otp_code,
                })

                # Send OTP via email
                email = EmailMessage(
                    email_subject,
                    email_message,
                    settings.EMAIL_HOST_USER,
                    [myuser.email],
                )
                email.content_subtype = 'html'  # Ensure the email is sent as HTML
                email.send(fail_silently=True)

                messages.success(request, "تم إرسال رمز التحقق إلى بريدك الإلكتروني.")
                request.session['username'] = myuser.username
                return redirect('verify_activation_otp')
            else:
                messages.error(request, "الحساب مفعل لهذا البريد الإلكتروني.")
                return redirect('resend_activation_email_view')

        except ObjectDoesNotExist:
            messages.error(request, "البريد الإلكتروني غير موجود.")
            return redirect('resend_activation_email_view')

    else:
        # Check if the user is already authenticated
        if request.user.is_authenticated:
            return redirect('home')

    return render(request, "auth/resend_activation_email.html")


def verify_activation_otp_view(request):
    if request.method == 'POST':
        if 'resend_otp' in request.POST:  # Handle OTP resend request
            username = request.session.get('username')

            try:
                user = CustomUser.objects.get(username=username)

                if user.is_active:
                    messages.warning(request, "الحساب مفعل بالفعل.")
                    return redirect('login')

                # Check if the OTP was recently sent (cooldown period of 1 minute)
                recent_otp = Authentication_OTP.objects.filter(
                    created_by=user,
                    purpose="Account Activation"
                ).latest('created_at')

                # Compare the current time with the OTP's created_at time
                if recent_otp.created_at + timedelta(minutes=1) > now():
                    messages.error(request, "يمكنك إعادة إرسال رمز التحقق بعد دقيقة واحدة.")
                    return redirect('verify_activation_otp')

                # Generate a new OTP
                otp_code = generate_otp()
                expiry_time = now() + timedelta(minutes=5)

                # Get the user's IP address
                ip_address = request.META.get('REMOTE_ADDR', None)

                # Save the OTP to the database
                Authentication_OTP.objects.create(
                    otp_code=otp_code,
                    expiry_time=expiry_time,
                    created_by=user,
                    purpose="Account Activation",
                    ip_address=ip_address,
                )

                # Render the OTP email template
                email_subject = "رمز التحقق لتفعيل الحساب (إعادة إرسال)"
                email_message = render_to_string('auth/otp_template.html', {
                    'name': user.first_name,
                    'otp_code': otp_code,
                })

                # Send OTP via email
                email = EmailMessage(
                    email_subject,
                    email_message,
                    settings.EMAIL_HOST_USER,
                    [user.email],
                )
                email.content_subtype = 'html'  # Ensure the email is sent as HTML
                email.send(fail_silently=True)

                messages.success(request, "تم إعادة إرسال رمز التحقق إلى بريدك الإلكتروني.")
                return redirect('verify_activation_otp')

            except CustomUser.DoesNotExist:
                messages.error(request, "المستخدم غير موجود.")
                return redirect('resend_activation_email_view')

        else:  # Handle OTP verification
            otp_code = request.POST.get('otp')
            username = request.session.get('username')

            try:
                user = CustomUser.objects.get(username=username)

                # Validate OTP
                otp_record = Authentication_OTP.objects.filter(
                    created_by=user,
                    purpose="Account Activation",
                    otp_code=otp_code,
                    expiry_time__gt=now(),  # Ensure OTP is not expired
                ).first()

                if otp_record:
                    # Mark the OTP as used
                    otp_record.is_used = True
                    otp_record.save()

                    # Activate the user's account
                    user.is_active = True
                    user.save()

                    messages.success(request, "تم تفعيل حسابك بنجاح.")
                    return redirect('login')  # Change 'login' to your desired URL
                else:
                    messages.error(request, "رمز التحقق غير صحيح أو منتهي الصلاحية.")
                    return redirect('verify_activation_otp')

            except CustomUser.DoesNotExist:
                messages.error(request, "المستخدم غير موجود.")
                return redirect('resend_activation_email_view')

    return render(request, "auth/verify_activation_otp.html")


def signin(request):
    if request.method == 'POST':
        identifier = request.POST.get("username")  # Could be email, national ID, or username
        password = request.POST.get("password", None)

        # Handle email, national ID, or username
        user = CustomUser.objects.filter(
            email=identifier
        ).first() or CustomUser.objects.filter(
            national_id=identifier
        ).first() or CustomUser.objects.filter(
            username=identifier
        ).first()

        if user and user.is_active:
            user_auth = authenticate(username=user.username, password=password)
            if user_auth:
                login(request, user_auth)
                return redirect("home")
            else:
                messages.error(request, "كلمة المرور أو اسم المستخدم خطأ!")
        else:
            messages.error(request, "المستخدم غير موجود أو الحساب غير نشط.")

    return render(request, "registration/login.html")


def phone_login_page(request):
    """
    Render the phone login page.
    """
    return render(request, 'registration/phone_login.html')


def send_otp_phone_login(request):
    """
    Handle OTP sending functionality for phone login using fetch.
    """
    if request.method == 'POST':
        phone_number = request.POST.get('phonenumber')

        print("phone number is " + phone_number)

        # Validate if the phone number exists in the database
        user = CustomUser.objects.filter(phonenumber=phone_number).first()
        if user:
            otp_code = generate_otp()

            print(otp_code)
            
            # Save OTP in the database
            Authentication_OTP.objects.create(
                otp_code=otp_code,
                expiry_time=now() + timedelta(minutes=5),
                created_by=user,
                purpose="Login by Phone number",
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:512],
            )
            
            # Send OTP via SMS
            if send_otp_via_sms(phone_number, otp_code, user):
                print("SMS is sent.")
                return JsonResponse({'success': True, 'message': "تم إرسال رمز التحقق إلى رقم الجوال.", 'phone_number': phone_number})
            else:
                return JsonResponse({'success': False, 'message': "فشل في إرسال رمز التحقق. حاول مرة أخرى."}, status=500)
        else:
            return JsonResponse({'success': False, 'message': "رقم الجوال غير مسجل."}, status=404)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)


def verify_otp_login(request):
    if request.method == 'POST':
        phone_number = request.POST.get("phone_number")
        otp_code = request.POST.get("otp", "").strip()

        user = CustomUser.objects.filter(phonenumber=phone_number).first()
        if user:
            otp_record = Authentication_OTP.objects.filter(
                created_by=user,
                otp_code=otp_code,
                is_used=False,
                expiry_time__gt=now(),
                purpose="Login by Phone number",
            ).order_by('-created_at').first()
            
            if otp_record:
                # Mark OTP as used
                otp_record.is_used = True
                otp_record.used_at = now()
                otp_record.save()

                user.save()

                # Log the user in
                login(request, user)
                messages.success(request, "تم التحقق من رقم هاتفك بنجاح!", 200)
                return redirect("home")
            else:
                messages.error(request, "رمز التحقق غير صحيح أو منتهي الصلاحية.")
                request.session['phonenumber'] = phone_number
                return redirect("phone_login_page")
        else:
            messages.error(request, "رقم الهاتف غير مسجل.", status=404)
            request.session['phonenumber'] = phone_number
            return redirect("phone_login_page")


@login_required(login_url="/login")
def logout_user(request):

    logout(request)
    return redirect("home")


def password_reset_request(request):
    if request.method == 'POST':
        reset_method = request.POST.get("reset_method", None)
        identifier = request.POST.get("identifier", None)

        if reset_method and identifier:
            try:
                # Fetch user based on the selected reset method
                if reset_method == "email":
                    user = CustomUser.objects.get(Q(email=identifier) | Q(national_id=identifier))
                    
                    # Generate OTP and save in database
                    otp_code = generate_otp()
                    expiry_time = now() + timedelta(minutes=5)  # OTP expires in 5 minutes
                    Authentication_OTP.objects.create(
                        created_by=user,
                        otp_code=otp_code,
                        purpose="Password Reset",
                        expiry_time=expiry_time
                    )
                    
                    # Send OTP via email
                    send_otp_email(user, otp_code)
                    
                    # Store user ID in session for verification
                    request.session['otp_user_id'] = user.id
                    
                    messages.success(request, "تم إرسال رمز التحقق إلى بريدك الإلكتروني.")
                    return redirect('password_reset_done')

                elif reset_method == "phone":
                    user = CustomUser.objects.get(phonenumber=identifier)  # Assuming `phonenumber` is unique
                    
                    # Generate OTP and save in database
                    otp_code = generate_otp()
                    expiry_time = now() + timedelta(minutes=5)  # OTP expires in 5 minutes
                    Authentication_OTP.objects.create(
                        created_by=user,
                        otp_code=otp_code,
                        purpose="Password Reset",
                        expiry_time=expiry_time
                    )
                    
                    # Send OTP via SMS (replace with your SMS logic)
                    sms_message = f"استخدام الرمز التالي لإعادة تعيين كلمة المرور: {otp_code}"
                    if send_otp_via_sms(user.phonenumber, sms_message, user):
                        request.session['otp_user_id'] = user.id
                        messages.success(request, "تم إرسال رمز التحقق إلى هاتفك.")
                        return redirect('password_reset_done')
                    else:
                        messages.error(request, "تعذر إرسال الرسالة. حاول مرة أخرى.")
                        return redirect("password_reset")

            except CustomUser.DoesNotExist:
                messages.error(request, "لم يتم العثور على بيانات مطابقة.")
                return redirect("password_reset")
        else:
            messages.error(request, "يرجى اختيار طريقة الإدخال الصحيحة.")
            return redirect("password_reset")

    return render(request, "auth/password_reset_form.html")


def password_reset_done(request):
    """
    Render the password reset done page.
    """
    if request.method == 'POST':
        otp_code = request.POST.get('otp_code', "").strip()
        user_id = request.session.get('otp_user_id')

        if user_id and otp_code:
            user = get_object_or_404(CustomUser, id=user_id)

            # Retrieve OTP record from the database
            otp_record = Authentication_OTP.objects.filter(
                created_by=user,
                otp_code=otp_code,
                is_used=False,
                expiry_time__gt=now(),
                purpose="Password Reset",
            ).order_by('-created_at').first()

            if otp_record:
                # Mark OTP as used
                otp_record.is_used = True
                otp_record.used_at = now()
                otp_record.save()

                # Redirect to password reset confirmation page
                del request.session['otp_user_id']  # Remove user ID from session
                messages.success(request, "تم التحقق من الرمز بنجاح.")
                return redirect('password_reset_confirm', 
                                uidb64=urlsafe_base64_encode(force_bytes(user.pk)),
                                token=default_token_generator.make_token(user))
            else:
                messages.error(request, "رمز التحقق غير صحيح أو منتهي الصلاحية.")
        else:
            messages.error(request, "حدث خطأ. يرجى المحاولة مرة أخرى.")
    else:
        return render(request, 'auth/password_reset_done.html')


def validate_email(request):

    email = request.POST.get('email', None)

    if email is None:
        return HttpResponse("true")
    else:
        data = "false"

        cu_data = not CustomUser.objects.filter(email__iexact=email).exists()
        s_data = not Supporter.objects.filter(email__iexact=email).exists()

        if cu_data and s_data:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)


def validate_username(request):

    username = request.POST.get('username', None)

    if username is None:
        return HttpResponse("true")
    else:

        data = not CustomUser.objects.filter(username__exact=username).exists()
        if data is True:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)


def validate_phonenumber(request):

    phone_number = request.POST.get('phonenumber', None)

    if phone_number is None:
        return HttpResponse("true")
    else:
        data = "false"

        cu_data = not CustomUser.objects.filter(
            phonenumber=phone_number).exists()
        s_data = not Supporter.objects.filter(
            phone_number=phone_number).exists()

        if cu_data and s_data:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)


@group_required("Management")
@login_required(login_url="/login")
def dashboard(request):
    beneficiaries_num = beneficiary.objects.count()
    dependents_num = dependent.objects.count()

    # Example: Counting beneficiaries by specific categories
    orphans_num = beneficiary.objects.filter(category='أسرة أيتام').count()
    widows_num = beneficiary.objects.filter(category='أسرة أرملة').count()

    # Example: Beneficiaries with medical needs
    beneficiaries_with_medical_needs = beneficiary.objects.filter(
        health_status='غير جيدة').count()

    # Example: Beneficiaries with financial needs
    beneficiaries_with_financial_needs = beneficiary.objects.filter(
        work_status='لا').count()

    # Calculating total income and expenses for all beneficiaries
    total_income = beneficiary_income_expense.objects.aggregate(
        total=Sum(F('salary_in') + F('social_insurance_in') + F('charity_in') + F('social_warranty_in') +
                  F('pension_agency_in') + F('citizen_account_in') + F('benefactor_in') + F('other_in'))
    )['total']

    total_expense = beneficiary_income_expense.objects.aggregate(
        total=Sum(F('housing_rent_ex') + F('electricity_bills_ex') + F('water_bills_ex') + F('transportation_ex') +
                  F('health_supplies_ex') + F('food_supplies_ex') + F('educational_supplies_ex') + F('proven_debts_ex') + F('other_ex'))
    )['total']
    # Calculating average income and expenses
    num_beneficiaries_with_income_expense = beneficiary_income_expense.objects.count()
    average_income = round(total_income / num_beneficiaries_with_income_expense, 2) if num_beneficiaries_with_income_expense else 0
    average_expense = round(total_expense / num_beneficiaries_with_income_expense, 2) if num_beneficiaries_with_income_expense else 0

    context = {
        'beneficiaries_num': beneficiaries_num,
        'dependents_num': dependents_num,
        'orphans_num': orphans_num,
        'widows_num': widows_num,
        'beneficiaries_with_medical_needs': beneficiaries_with_medical_needs,
        'beneficiaries_with_financial_needs': beneficiaries_with_financial_needs,
        'average_income': average_income,
        'average_expense': average_expense,
    }

    return render(request, "dashboard/dashboard_home.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_supporters_requests(request):

    context = {}

    try:
        # List of supporters requests
        supporters_request_list = Supporter_request.objects.all()

        # List of beneficiaries
        beneficiary_list = beneficiary.objects.all()

        # List of supporters
        supporters_list = Supporter.objects.all()

        context = {
            'beneficiaries': beneficiary_list,
            'supporters': supporters_list,
            'supporters_requests': supporters_request_list,
        }
    except ObjectDoesNotExist:
        print("[Error] - Object does not exist, dashboard_supporters_request.")
        return JsonResponse({"message": "Object does not exist!", "code": "404"})

    return render(request, "dashboard/supporters_requests.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_beneficiaries_requests(request):
    # Fetch search and pagination parameters from DataTables
    search_value = request.GET.get('search[value]', '')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    draw = int(request.GET.get('draw', 1))

    # Retrieve order parameters
    order_column_index = request.GET.get('order[0][column]', '0')
    order_direction = request.GET.get('order[0][dir]', 'asc')

    # Map DataTables column index to model field names
    columns = [
        'id', 'beneficiary__first_name', 'beneficiary__last_name',
        'request_type', 'status', 'created_at', 'reviewed_at', 'reviewed_by__username'
    ]
    order_column = columns[int(order_column_index)]
    if order_direction == 'desc':
        order_column = f'-{order_column}'

    # Initial queryset with optional search filter
    beneficiary_requests = Beneficiary_request.objects.prefetch_related('beneficiary').order_by(order_column)

    # Determine if search_value is numeric or contains digits
    if search_value.isdigit(): # Handles numeric strings
        beneficiary_requests = beneficiary_requests.filter(
            Q(user__first_name__icontains=search_value) |
            Q(user__last_name__icontains=search_value) |
            Q(request_type__icontains=search_value) |
            Q(status__icontains=search_value) |
            Q(id=search_value)
        )
    elif any(char.isdigit() for char in search_value):  # Search contains both digits and characters
        beneficiary_requests = beneficiary_requests.filter(
            Q(user__first_name__icontains=search_value) |
            Q(user__last_name__icontains=search_value) |
            Q(request_type__icontains=search_value) |
            Q(status__icontains=search_value)
        )
    else: # No numeric characters, search only other fields
        beneficiary_requests = beneficiary_requests.filter(
            Q(user__first_name__icontains=search_value) |
            Q(user__last_name__icontains=search_value) |
            Q(request_type__icontains=search_value) |
            Q(status__icontains=search_value)
        )

    # Pagination logic
    paginator = Paginator(beneficiary_requests, length)
    page_number = start // length + 1
    page_obj = paginator.get_page(page_number)
    
    # Check if the request is AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        data = [
            {
                'id': req.id,
                'first_name': req.user.first_name if req.user.first_name else "----",
                'last_name': req.user.last_name if req.user.last_name else "----",
                'request_type': req.request_type,
                'status': req.status,
                'created_at': req.created_at.strftime('%H:%M:%S -- %Y-%m-%d '),  # Format date as needed
                'reviewed_at': req.reviewed_at.strftime('%H:%M:%S -- %Y-%m-%d ') if req.reviewed_at else "----",
                # Serialize only the `username` or relevant field from `reviewed_by`
                'reviewed_by': req.reviewed_by.username if req.reviewed_by else "----",
                'details_url': f"/dashboard/beneficiaries/{req.beneficiary.id}/requests/{req.id}"
            }
            for req in page_obj
        ]
        # Return paginated response with DataTables format
        return JsonResponse({
            'draw': draw,
            'recordsTotal': paginator.count,
            'recordsFiltered': paginator.count if not search_value else beneficiary_requests.count(),
            'data': data,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number
        })
    # Pass the first page of results to the template for initial load
    context = {
        "beneficiary_requests": page_obj,
        "beneficiary_request_headers": ['رقم الطلب', 'نوع الطلب', 'الحالة', 'تاريخ الإرسال', 'مُراجع الطلب', 'الملاحظات', 'الإجراءات'],
    }
    return render(request, "dashboard/beneficiaries_requests.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_reports(request):

    if request.method == "POST":

        # Get beneficiary table data (all)
        beneficiary_arr = beneficiary.objects.all()

        # Retrive form data
        beneficiary_first_name = request.POST.get("beneficiary_first_name")
        beneficiary_last_name = request.POST.get("beneficiary_last_name")
        national_id = request.POST.get("beneficiary_national_id")
        category = request.POST.get("beneficiary_category")
        marital_status = request.POST.get("beneficiary_marital_status")
        is_qualified = request.POST.get("beneficiary_is_qualified")

        # Passing the form data to the session data
        request.session["beneficiary_first_name"] = beneficiary_first_name
        request.session["beneficiary_last_name"] = beneficiary_last_name
        request.session["beneficiary_national_id"] = national_id
        request.session["beneficiary_category"] = category
        request.session["beneficiary_marital_status"] = marital_status
        request.session["beneficiary_is_qualified"] = is_qualified

        # Validate query param
        if is_valid_queryparam(beneficiary_first_name, type=1):
            beneficiary_arr = beneficiary_arr.filter(
                first_name__icontains=beneficiary_first_name)

        if is_valid_queryparam(beneficiary_last_name, type=1):
            beneficiary_arr = beneficiary_arr.filter(
                last_name__icontains=beneficiary_last_name)

        if is_valid_queryparam(national_id, type=1):
            beneficiary_arr = beneficiary_arr.filter(national_id=national_id)

        if is_valid_queryparam(category, type=2):
            beneficiary_arr = beneficiary_arr.filter(category=category)

        if is_valid_queryparam(marital_status, type=2):
            beneficiary_arr = beneficiary_arr.filter(
                marital_status=marital_status)

        # Keep the original value to be sent back in the response
        is_qualified_val = is_qualified

        if is_valid_queryparam(is_qualified, type=2):
            if is_qualified == "مؤهل":
                is_qualified = True
            else:
                is_qualified = False

            beneficiary_arr = beneficiary_arr.filter(is_qualified=is_qualified)

        # Prepare pagination
        paginator = Paginator(beneficiary_arr, IPP_DASHBOARD_REPORTS)
        page_number = request.GET.get('page')
        beneficiary_arr = paginator.get_page(page_number)

        context = {
            "beneficiaries_headers": [
                "رقم الملف",
                "الأسم الأول",
                "الأسم الأخير",
                "رقم الهوية",
                "التصنيف",
                "الحالة الاجتماعية",
                "مؤهل؟"
            ],
            "beneficiaries": beneficiary_arr,
            "first_name": beneficiary_first_name,
            "last_name": beneficiary_last_name,
            "national_id": national_id,
            "category": category,
            "marital_status": marital_status,
            "is_qualified": is_qualified_val,
        }

        return render(request, "dashboard/generate_reports.html", context)

    else:
        return render(request, "dashboard/generate_reports.html")


@group_required("Management")
@login_required(login_url="/login")
def dashboard_reports_post(request):

    if request.method == "GET":

        # Get beneficiary table data (all)
        beneficiary_arr = beneficiary.objects.all()

        # Retrive form data
        beneficiary_first_name = request.GET.get("beneficiary_first_name")
        beneficiary_last_name = request.GET.get("beneficiary_last_name")
        national_id = request.GET.get("beneficiary_national_id")
        category = request.GET.get("beneficiary_category")
        marital_status = request.GET.get("beneficiary_marital_status")
        is_qualified = request.GET.get("beneficiary_is_qualified")

        # Passing the form data to the session data
        request.session["beneficiary_first_name"] = beneficiary_first_name
        request.session["beneficiary_last_name"] = beneficiary_last_name
        request.session["beneficiary_national_id"] = national_id
        request.session["beneficiary_category"] = category
        request.session["beneficiary_marital_status"] = marital_status
        request.session["beneficiary_is_qualified"] = is_qualified

        # Validate query param
        if is_valid_queryparam(beneficiary_first_name, type=1):
            beneficiary_arr = beneficiary_arr.filter(
                first_name__icontains=beneficiary_first_name)

        if is_valid_queryparam(beneficiary_last_name, type=1):
            beneficiary_arr = beneficiary_arr.filter(
                last_name__icontains=beneficiary_last_name)

        if is_valid_queryparam(national_id, type=1):
            beneficiary_arr = beneficiary_arr.filter(national_id=national_id)

        if is_valid_queryparam(category, type=2):
            beneficiary_arr = beneficiary_arr.filter(category=category)

        if is_valid_queryparam(marital_status, type=2):
            beneficiary_arr = beneficiary_arr.filter(
                marital_status=marital_status)

        # Keep the original value to be sent back in the response
        is_qualified_val = is_qualified

        if is_valid_queryparam(is_qualified, type=2):
            if is_qualified == "مؤهل":
                is_qualified = True
            else:
                is_qualified = False

            beneficiary_arr = beneficiary_arr.filter(is_qualified=is_qualified)

        # Prepare pagination
        paginator = Paginator(beneficiary_arr, IPP_DASHBOARD_REPORTS)
        page_number = request.GET.get('page')
        beneficiary_arr = paginator.get_page(page_number)

        context = {
            "beneficiaries_headers": [
                "رقم الملف",
                "الأسم الأول",
                "الأسم الأخير",
                "رقم الهوية",
                "التصنيف",
                "الحالة الاجتماعية",
                "مؤهل؟"
            ],
            "beneficiaries": beneficiary_arr,
            "first_name": beneficiary_first_name,
            "last_name": beneficiary_last_name,
            "national_id": national_id,
            "category": category,
            "marital_status": marital_status,
            "is_qualified": is_qualified_val,
        }

        return render(request, "dashboard/generate_reports.html", context)

    else:
        return render(request, "dashboard/generate_reports.html")


@group_required("Management")
@login_required(login_url="/login")
def export_excel(request):

    # Get beneficiary table data (all)
    beneficiary_arr = beneficiary.objects.all()

    # Ensure data is in the session (request.session is used to retrieve the data included in the session)
    if 'beneficiary_first_name' in request.session:
        beneficiary_first_name = request.session["beneficiary_first_name"]
    else:
        beneficiary_first_name = None

    if 'beneficiary_last_name' in request.session:
        beneficiary_last_name = request.session["beneficiary_last_name"]
    else:
        beneficiary_last_name = None

    if 'beneficiary_national_id' in request.session:
        beneficiary_national_id = request.session["beneficiary_national_id"]
    else:
        beneficiary_national_id = None

    if 'beneficiary_category' in request.session:
        beneficiary_category = request.session["beneficiary_category"]
    else:
        beneficiary_category = None

    if 'beneficiary_marital_status' in request.session:
        beneficiary_marital_status = request.session["beneficiary_marital_status"]
    else:
        beneficiary_marital_status = None

    if 'beneficiary_is_qualified' in request.session:
        beneficiary_is_qualified = request.session["beneficiary_is_qualified"]
    else:
        beneficiary_is_qualified = None

    # Validate query param
    if is_valid_queryparam(beneficiary_first_name, type=1):
        beneficiary_arr = beneficiary_arr.filter(
            first_name__icontains=beneficiary_first_name)

    if is_valid_queryparam(beneficiary_last_name, type=1):
        beneficiary_arr = beneficiary_arr.filter(
            last_name__icontains=beneficiary_last_name)

    if is_valid_queryparam(beneficiary_national_id, type=1):
        beneficiary_arr = beneficiary_arr.filter(
            national_id=beneficiary_national_id)

    if is_valid_queryparam(beneficiary_category, type=2):
        beneficiary_arr = beneficiary_arr.filter(category=beneficiary_category)

    if is_valid_queryparam(beneficiary_marital_status, type=2):
        beneficiary_arr = beneficiary_arr.filter(
            marital_status=beneficiary_marital_status)

    if is_valid_queryparam(beneficiary_is_qualified, type=2):
        if beneficiary_marital_status == "مؤهل":
            beneficiary_marital_status = True
        else:
            beneficiary_marital_status = False

        beneficiary_arr = beneficiary_arr.filter(
            is_qualified=beneficiary_marital_status)

    if beneficiary_first_name is None or beneficiary_first_name == '':
        beneficiary_first_name = "الكل"
    else:
        beneficiary_first_name = beneficiary_first_name

    if beneficiary_last_name is None or beneficiary_last_name == '':
        beneficiary_last_name = "الكل"
    else:
        beneficiary_last_name = beneficiary_last_name

    if beneficiary_national_id is None or beneficiary_national_id == '':
        beneficiary_national_id = "الكل"
    else:
        beneficiary_national_id = beneficiary_national_id

    if beneficiary_category is None or beneficiary_category == "اختار...":
        beneficiary_category = "الكل"
    else:
        beneficiary_category = beneficiary_category

    if beneficiary_marital_status is None or beneficiary_marital_status == "اختار...":
        beneficiary_marital_status = "الكل"
    else:
        beneficiary_marital_status = beneficiary_marital_status

    if beneficiary_is_qualified is None or beneficiary_is_qualified == "اختار...":
        beneficiary_is_qualified = "الكل"
    else:
        beneficiary_is_qualified = beneficiary_is_qualified

    # Let the browser know what type of file is included in the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Name the file
    response['Content-Disposition'] = 'attachment; filename="beneficiary' + \
        str(datetime.now()) + '.xlsx"'

    # Workbook object
    workbook = Workbook()

    worksheet = workbook.active

    # Merge the first six rows which indicate the type of data included
    worksheet.merge_cells('A1:H1')
    worksheet.merge_cells('A2:H2')
    worksheet.merge_cells('A3:H3')
    worksheet.merge_cells('A4:H4')
    worksheet.merge_cells('A5:H5')
    worksheet.merge_cells('A6:H6')

    # Style the first row
    first_cell = worksheet['A1']
    first_cell.value = "الأسم الأول: " + " " + beneficiary_first_name
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the second row
    second_cell = worksheet['A2']
    second_cell.value = "الأسم الأخير: " + " " + beneficiary_last_name
    second_cell.font = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the third row
    third_cell = worksheet['A3']
    third_cell.value = "رقم الهوية: " + " " + beneficiary_national_id
    third_cell.font = Font(bold=True, color="246ba1")
    third_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the forth row
    forth_cell = worksheet['A4']
    forth_cell.value = "التصنيف: " + " " + beneficiary_category
    forth_cell.font = Font(bold=True, color="246ba1")
    forth_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the fifth row
    fifth_cell = worksheet['A5']
    fifth_cell.value = "الحالة الاجتماعية: " + " " + beneficiary_marital_status
    fifth_cell.font = Font(bold=True, color="246ba1")
    fifth_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the sixth row
    sixth_cell = worksheet['A6']
    sixth_cell.value = "مؤهل أم لا: " + " " + beneficiary_is_qualified
    sixth_cell.font = Font(bold=True, color="246ba1")
    sixth_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'AA'

    # Define the titles for columns
    columns = ['#', 'رقم الملف', 'الأسم الأول',
               'الأسم الأخير', 'رقم الهوية', 'التصنيف', 'الحالة الاجتماعية', 'مؤهل؟']
    row_num = 7

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['H7']
        seventh_cell.alignment = Alignment(horizontal="right")

    for beneficiaries in beneficiary_arr:
        row_num += 1

        # Define the data for each cell in the row
        row = [beneficiaries.id, beneficiaries.file_no, beneficiaries.first_name,
               beneficiaries.last_name, beneficiaries.national_id, beneficiaries.category, beneficiaries.marital_status, beneficiaries.is_qualified]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            # if isinstance(cell_value, decimal.Decimal):
            #     cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    workbook.save(response)
    return response
# This is for demonstration purposes only. In production, use CSRF protection.


@csrf_exempt
@login_required(login_url="/login")
def beneficiary_indiv(request, user_id):

    if request.method == 'POST':
        data = request.POST
        files = request.FILES

        # Get all the attachments of beneficiary
        national_id_file = files.get('fileBeneficiaryNationalID', None)
        national_address_file = files.get(
            'fileBeneficiaryNationalAddress', None)
        dept_instrument_file = files.getlist('fileDeptInstrument')
        pension_social_insurance_file = files.getlist(
            'filePensionOrSocialInsuranceInquiry')
        father_husband_death_certificate_file = files.get(
            'fileFatherOrHusbandDeathCertificate', None)
        letter_from_prison_file = files.getlist('fileLetterFromPrison')
        divorce_deed_file = files.get('fileDivorceDeed', None)
        children_responsibility_deed_file = files.getlist(
            'fileChildrenResponsibilityDeed')
        other_files = files.getlist('fileOther')
        lease_contract_or_title_deed_file = files.getlist(
            'fileLeaseContractOrTitleDeed')
        water_or_electricity_bills_file = files.getlist(
            'fileWaterOrElectricityBills')
        dependent_national_id_file = files.getlist(
            'fileNationalIDForBeneficiaryDependents')
        social_warranty_inquiry_file = files.getlist(
            'fileSocialWarrantyInquiry')

        # Accessing the data for beneficiary
        category = data.get('personalinfo_category', None)
        marital_status = data.get('personalinfo_marital_status', None)
        educational_level = data.get('personalinfo_educational_level', None)
        date_of_death_of_father_or_husband = data.get(
            'personalinfo_date_of_death_of_father_or_husband', None)
        if date_of_death_of_father_or_husband is not None:
            date_of_death_of_father_or_husband = convert_to_date(
                date_of_death_of_father_or_husband)
        washing_place = data.get('personalinfo_washing_place', None)
        health_status = data.get('personalinfo_health_status', None)
        disease_type = data.get('personalinfo_disease_type', None)
        work_status = data.get('personalinfo_work_status', None)
        employer = data.get('personalinfo_employer', None)
        bank_type = data.get('beneficiaryinfo_bank', None)
        bank_iban = data.get('beneficiaryinfo_iban', None)
        family_issues = data.get('familyinfo_family_issues', None)
        family_needs = data.get('familyinfo_needs_type', None)

        beneficiary_obj = beneficiary(
            category=category,
            marital_status=marital_status,
            educational_level=educational_level,
            death_date_father_husband=date_of_death_of_father_or_husband,
            washing_place=washing_place,
            health_status=health_status,
            disease_type=disease_type,
            work_status=work_status,
            employer=employer,
            bank_type=bank_type,
            bank_iban=bank_iban,
            family_issues=family_issues,
            family_needs=family_needs,
            user=request.user,  # Pass user object to link it to this beneficiary profile
        )
        beneficiary_obj.save(category_seg="CAT", region_seg="SA")

        # Store attachments of beneficiary -------------------
        # Store all file objects in a list
        file_list = []

        # Create beneficiary attachment for "national id"
        if national_id_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="national_id",
                file_object=national_id_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "national address"
        if national_address_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="national_address",
                file_object=national_address_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "dept instrument"
        for file_obj in dept_instrument_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="dept_instrument",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "pension or social insurance"
        for file_obj in pension_social_insurance_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="pension_social_insurance",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "father or husband death certificate"
        if father_husband_death_certificate_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="father_husband_death_cert",
                file_object=father_husband_death_certificate_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "letter from prison"
        for file_obj in letter_from_prison_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="letter_from_prison",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "Divorce Deed"
        if divorce_deed_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="divorce_deed",
                file_object=divorce_deed_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "children responsibility deed"
        for file_obj in children_responsibility_deed_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="children_responsibility_deed",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "other files"
        for file_obj in other_files:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="other_files",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "lease contract or title deed"
        for file_obj in lease_contract_or_title_deed_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="lease_contract_title_deed",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "water or electricity bills"
        for file_obj in water_or_electricity_bills_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="water_or_electricity_bills",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "dependent national id"
        for file_obj in dependent_national_id_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="dependent_national_id",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "social warranty inquiry"
        for file_obj in social_warranty_inquiry_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="social_warranty_inquiry",
                file_object=file_obj
            ))

        # instead of creating and saving each file separately, store them in a list, and save them all at once.
        if file_list:
            Beneficiary_attachment.objects.bulk_create(file_list)

        # Accessing the data for beneficiary_house -----------------
        building_number = data.get('houseinfo_building_number', None)
        street_name = data.get('houseinfo_street_name', None)
        neighborhood = data.get('houseinfo_neighborhood', None)
        city = data.get('houseinfo_city', None)
        postal_code = data.get('houseinfo_postal_code', None)
        additional_number = data.get('houseinfo_additional_number', None)
        unit = data.get('houseinfo_unit', None)
        location_url = data.get('houseinfo_location_url', None)
        housing_type = data.get('houseinfo_housing_type', None)
        housing_ownership = data.get('houseinfo_housing_ownership', None)

        beneficiary_house_obj = beneficiary_house(
            building_number=building_number,
            street_name=street_name,
            neighborhood=neighborhood,
            city=city,
            postal_code=postal_code,
            additional_number=additional_number,
            unit=unit,
            location_url=location_url,
            housing_type=housing_type,
            housing_ownership=housing_ownership,
            beneficiary_id=beneficiary_obj,
        )
        beneficiary_house_obj.save()

        # Accessing the data for beneficiary_income_expense, and converting str into float type
        salary_in = float(data.get('incomeinfo_salary', None))
        social_insurance_in = float(
            data.get('incomeinfo_social_insurance', None))
        charity_in = float(data.get('incomeinfo_charity', None))
        social_warranty_in = float(
            data.get('incomeinfo_social_warranty', None))
        pension_agency_in = float(data.get('incomeinfo_pension_agency', None))
        citizen_account_in = float(
            data.get('incomeinfo_citizen_account', None))
        benefactor_in = float(data.get('incomeinfo_benefactor', None))
        other_in = float(data.get('incomeinfo_other', None))
        housing_rent_ex = float(data.get('expensesinfo_housing_rent', None))
        electricity_bills_ex = float(
            data.get('expensesinfo_electricity_bills', None))
        water_bills_ex = float(data.get('expensesinfo_water_bills', None))
        transportation_ex = float(
            data.get('expensesinfo_transportation', None))
        health_supplies_ex = float(
            data.get('expensesinfo_health_supplies', None))
        food_supplies_ex = float(data.get('expensesinfo_food_supplies', None))
        educational_supplies_ex = float(data.get(
            'expensesinfo_educational_supplies', None))
        proven_debts_ex = float(data.get('expensesinfo_proven_debts', None))
        other_ex = float(data.get('expensesinfo_other', None))

        beneficiary_income_expense_obj = beneficiary_income_expense(
            salary_in=salary_in,
            social_insurance_in=social_insurance_in,
            charity_in=charity_in,
            social_warranty_in=social_warranty_in,
            pension_agency_in=pension_agency_in,
            citizen_account_in=citizen_account_in,
            benefactor_in=benefactor_in,
            other_in=other_in,
            housing_rent_ex=housing_rent_ex,
            electricity_bills_ex=electricity_bills_ex,
            water_bills_ex=water_bills_ex,
            transportation_ex=transportation_ex,
            health_supplies_ex=health_supplies_ex,
            food_supplies_ex=food_supplies_ex,
            educational_supplies_ex=educational_supplies_ex,
            proven_debts_ex=proven_debts_ex,
            other_ex=other_ex,
            beneficiary_id=beneficiary_obj,
        )
        beneficiary_income_expense_obj.save()

        # print("Beneficiary: ", first_name, second_name, last_name, date_of_birth, gender, national_id, national_id_exp_date, nationality, category, marital_status,
        #       educational_level, date_of_death_of_father_or_husband, washing_place, health_status, disease_type, work_status, employer, phone_number, email, bank_iban, bank_type, family_issues, family_needs)

        # print("\nBeneficiary House: ", building_number, street_name, neighborhood, city, postal_code,
        #       postal_code, additional_number, unit, location_url, housing_type, housing_ownership)

        # print("\nBeneficiary Income Expenses: ", salary_in, social_insurance_in, charity_in, social_warranty_in, pension_agency_in, citizen_account_in, benefactor_in, other_in,
        #       housing_rent_ex, electricity_bills_ex, water_bills_ex, transportation_ex, health_supplies_ex, food_supplies_ex, educational_supplies_ex, proven_debts_ex, other_ex)

        dependent_table = data.get('dependents-table', None)

        # Parse the JSON string into a Python object
        try:
            dependents_list = json.loads(dependent_table)
        except json.JSONDecodeError:
            print("Error parsing JSON")
            dependents_list = []

        # Now, you can iterate over the list of dependents

        for dep in dependents_list:
            # Extract the data for each field
            first_name = dep.get('firstName', '')
            second_name = dep.get('secondName', '')
            last_name = dep.get('lastName', '')
            gender = dep.get('gender', '')
            relationship = dep.get('relationship', '')
            educational_status = dep.get('educationalStatus', None)
            marital_status = dep.get('martialStatus', '')
            national_id = dep.get('nationalID', '')
            bank_iban = dep.get('bankIban', '')
            bank_type = dep.get('bankType', '')
            health_status = dep.get('healthStatus', None)
            needs_type = dep.get('needsType', '')
            educational_degree = dep.get('educationalDegree', '')
            dependent_date_of_birth_data = dep.get('dateOfBitrh', None)
            dependent_date_of_birth = None
            if dependent_date_of_birth_data is not None:
                dependent_date_of_birth = convert_to_date(
                    dependent_date_of_birth_data)
            national_id_exp_date = dep.get(
                'nationalIDExpDate', None)
            if national_id_exp_date is not None:
                national_id_exp_date = convert_to_date(national_id_exp_date)
            needs_description = dep.get('needsDescription', '')
            educational_level = dep.get('educationalLevel', None)
            disease_type = dep.get('diseaseType', None)
            dependent_income_table = json.loads(
                dep.get('dependentIncomeTable', []))

            work_status = dep.get('workStatus', None)
            employer = dep.get('employer', None)
            contribute_to_family_income = dep.get(
                'contributeToFamilyIncome', None)
            disability_check = dep.get('disabilityCheck', None)
            disability_type = dep.get('disabilityType', None)

            # Create a new dependent object and save it to the database
            new_dependent = dependent(
                first_name=first_name,
                second_name=second_name,
                last_name=last_name,
                gender=gender,
                relationship=relationship,
                date_of_birth=dependent_date_of_birth,
                national_id=national_id,
                national_id_exp_date=national_id_exp_date,
                bank_iban=bank_iban,
                bank_type=bank_type,
                marital_status=marital_status,
                educational_level=educational_level,
                educational_status=educational_status,
                health_status=health_status,
                disease_type=disease_type,
                needs_type=needs_type,
                educational_degree=educational_degree,
                needs_description=needs_description,
                work_status=work_status,
                employer=employer,
                contribute_to_family_income=contribute_to_family_income,
                disability_check=disability_check,
                disability_type=disability_type,
                beneficiary_id=beneficiary_obj
            )
            new_dependent.save()

            # Store list of income for dependent -------------------
            # Store all file objects in a list
            dependent_income_list = []

            for entry in dependent_income_table:
                # Extract the monthly income and remove commas
                income_amount_str = entry.get('income_amount', '')
                income_amount = Decimal(income_amount_str.replace(',', ''))
                income_source = entry.get('income_source', '')

                # Initialize dependent income list
                dependent_income_obj = Dependent_income(
                    source=income_source,
                    amount=income_amount,
                    dependent=new_dependent
                )
                dependent_income_list.append(dependent_income_obj)

            # Save dependent income objects
            if dependent_income_list:
                Dependent_income.objects.bulk_create(dependent_income_list)

        # Create Beneficiary_request object in the DB
        # Retrieve the object for the logged in user
        logged_in_user = request.user

        new_beneficiary_request = Beneficiary_request(
            user=logged_in_user,
            beneficiary=beneficiary_obj,
            status="انتظار",
            request_type="جديد",
        )
        new_beneficiary_request.save()

        # In case of successful submission and valid form data
        return JsonResponse({'redirect': '/confirmation', 'file_no': beneficiary_obj.file_no})

    elif request.method == 'GET':

        # Get the logged-in user
        logged_in_user = request.user

        context = {}
        try:
            # Retrieve the user whose profile is being requested
            user = CustomUser.objects.get(id=user_id)

            # Check if the logged-in user matches the requested user
            if logged_in_user != user:
                messages.error(request, "ليس لديك الصلاحية اللازمة!")
                return redirect('home')

            # Assuming `user_id` is the ID of the user you're checking for
            beneficiary_exists = beneficiary.objects.filter(user=user).exists()

            if beneficiary_exists:
                messages.error(
                    request, "لديك طلب مستفيد سابق! لا يمكنك طلب مستفيد أخر!")
                return redirect('home')
            else:

                return render(request, "main/beneficiary_form.html")
        except ObjectDoesNotExist:
            messages.error(request, "المستخدم غير موجود!")
            return redirect('home')

    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@login_required(login_url="/login")
def beneficiary_details(request, beneficiary_id):
    if request.method == 'GET':
        try:
            beneficiary_obj = beneficiary.objects.get(id=beneficiary_id)

            # Ensure this is working by printing the result below
            user_obj = CustomUser.objects.get(id=beneficiary_obj.user)

            beneficiary_housing_obj = beneficiary_house.objects.filter(
                beneficiary_id=beneficiary_id).first()

            housing_data = {}

            if beneficiary_housing_obj is not None:

                housing_data = {
                    'building_number': beneficiary_housing_obj.building_number,
                    'street_name': beneficiary_housing_obj.street_name,
                    'neighborhood': beneficiary_housing_obj.neighborhood,
                    'city': beneficiary_housing_obj.city,
                    'postal_code': beneficiary_housing_obj.postal_code,
                    'additional_number': beneficiary_housing_obj.additional_number,
                    'unit': beneficiary_housing_obj.unit,
                    'location_url': beneficiary_housing_obj.location_url,
                    'housing_type': beneficiary_housing_obj.housing_type,
                    'housing_ownership': beneficiary_housing_obj.housing_ownership
                }

            beneficiary_income_expense_obj = beneficiary_income_expense.objects.filter(
                beneficiary_id=beneficiary_id).first()

            income_expense_data = {}

            if beneficiary_income_expense_obj is not None:
                income_expense_data = {
                    'salary_in': beneficiary_income_expense_obj.salary_in,
                    'social_insurance_in': beneficiary_income_expense_obj.social_insurance_in,
                    'charity_in': beneficiary_income_expense_obj.charity_in,
                    'social_warranty_in': beneficiary_income_expense_obj.social_warranty_in,
                    'pension_agency_in': beneficiary_income_expense_obj.pension_agency_in,
                    'citizen_account_in': beneficiary_income_expense_obj.citizen_account_in,
                    'benefactor_in': beneficiary_income_expense_obj.benefactor_in,
                    'other_in': beneficiary_income_expense_obj.other_in,
                    'housing_rent_ex': beneficiary_income_expense_obj.housing_rent_ex,
                    'electricity_bills_ex': beneficiary_income_expense_obj.electricity_bills_ex,
                    'water_bills_ex': beneficiary_income_expense_obj.water_bills_ex,
                    'transportation_ex': beneficiary_income_expense_obj.transportation_ex,
                    'health_supplies_ex': beneficiary_income_expense_obj.health_supplies_ex,
                    'food_supplies_ex': beneficiary_income_expense_obj.food_supplies_ex,
                    'educational_supplies_ex': beneficiary_income_expense_obj.educational_supplies_ex,
                    'proven_debts_ex': beneficiary_income_expense_obj.proven_debts_ex,
                    'other_ex': beneficiary_income_expense_obj.other_ex
                }

            dependent_list = dependent.objects.filter(
                beneficiary_id=beneficiary_id).all()

            dependent_data = []

            for dependent_obj in dependent_list:

                # Initialize dependent income list with every dependent
                dependent_income_data = []

                # Retrieve the dependent income infomration
                dependent_income_list = Dependent_income.objects.filter(
                    dependent=dependent_obj).all()

                # Add the data into the dependent income list
                for dependent_income_obj in dependent_income_list:
                    dependent_income_data.append({
                        'income_source': dependent_income_obj.source,
                        'income_amount': dependent_income_obj.amount,
                    })

                dependent_data.append({
                    'dependent_id': dependent_obj.id,
                    'dependent_first_name': dependent_obj.first_name,
                    'dependent_second_name': dependent_obj.second_name,
                    'dependent_last_name': dependent_obj.last_name,
                    'dependent_gender': dependent_obj.gender,
                    'dependent_relationship': dependent_obj.relationship,
                    'dependent_educational_status': dependent_obj.educational_status,
                    'dependent_marital_status': dependent_obj.marital_status,
                    'dependent_national_id': dependent_obj.national_id,
                    'dependent_national_id_exp_date': dependent_obj.national_id_exp_date,
                    'dependent_bank_iban': dependent_obj.bank_iban,
                    'dependent_bank_type': dependent_obj.bank_type,
                    'dependent_health_status': dependent_obj.health_status,
                    'dependent_needs_type': dependent_obj.needs_type,
                    'dependent_educational_degree': dependent_obj.educational_degree,
                    'dependent_date_of_birth': dependent_obj.date_of_birth,
                    'dependent_needs_description': dependent_obj.needs_description,
                    'dependent_educational_level': dependent_obj.educational_level,
                    'dependent_disease_type': dependent_obj.disease_type,
                    'dependent_work_status': dependent_obj.work_status,
                    'dependent_employer': dependent_obj.employer,
                    'dependent_contribute_to_family_income': dependent_obj.contribute_to_family_income,
                    'dependent_disability_check': dependent_obj.disability_check,
                    'dependent_disability_type': dependent_obj.disability_type,
                    'dependent_income_data': dependent_income_data,
                })

            beneficiary_attachment_list = []

            attachments_list = Beneficiary_attachment.objects.filter(
                beneficiary_id=beneficiary_obj.id).all()

            for attachment in attachments_list:
                # A variable that holds the attachment type in Arabic
                attachment_type_ar = ""

                if attachment.file_type == "national_id":
                    attachment_type_ar = "صورة الهوية الوطنية/الإقامة"
                elif attachment.file_type == "national_address":
                    attachment_type_ar = "العنوان الوطني"
                elif attachment.file_type == "dept_instrument":
                    attachment_type_ar = "صك الدين"
                elif attachment.file_type == "pension_social_insurance":
                    attachment_type_ar = "مشهد التقاعد أو التأمينات الاجتماعية"
                elif attachment.file_type == "father_husband_death_cert":
                    attachment_type_ar = "شهادة الوفاة للزوج / الأب"
                elif attachment.file_type == "letter_from_prison":
                    attachment_type_ar = "خطاب من السجن"
                elif attachment.file_type == "divorce_deed":
                    attachment_type_ar = "صك الطلاق"
                elif attachment.file_type == "children_responsibility_deed":
                    attachment_type_ar = "صك إعالة الأبناء"
                elif attachment.file_type == "other_files":
                    attachment_type_ar = "مستندات أخرى"
                elif attachment.file_type == "lease_contract_title_deed":
                    attachment_type_ar = "عقد الإيجار الالكتروني من منصة إيجار أو صك ملكية"
                elif attachment.file_type == "water_or_electricity_bills":
                    attachment_type_ar = "الفواتير (كهرباء - ماء)"
                elif attachment.file_type == "dependent_national_id":
                    attachment_type_ar = "صورة الهوية الوطنية/الإقامة للمرافقين"
                elif attachment.file_type == "social_warranty_inquiry":
                    attachment_type_ar = "مشهد الضمان الاجتماعي"
                else:
                    attachment_type_ar = attachment.file_type

                beneficiary_attachment_list.append({
                    'file_path': attachment.file_object.url,
                    'file_extension': file_extension(attachment.file_object.url),
                    'file_name': attachment.filename().split(".")[0],
                    'file_size': attachment.file_size,
                    'attachment_type': attachment_type_ar,
                })
            # print("attachments: ", beneficiary_attachment_list)

            data = {
                'id': beneficiary_obj.id,
                'file_no': beneficiary_obj.file_no,
                'first_name': beneficiary_obj.first_name,
                'second_name': beneficiary_obj.second_name,
                'last_name': beneficiary_obj.last_name,
                'nationality': user_obj.nationality,
                'gender': user_obj.gender,
                'date_of_birth': user_obj.date_of_birth,
                'phone_number': user_obj.phonenumber,
                'email': user_obj.email,
                'national_id': user_obj.national_id,
                'national_id_exp_date': user_obj.national_id_exp_date,
                'is_qualified': beneficiary_obj.is_qualified,
                'category': beneficiary_obj.category,
                'marital_status': beneficiary_obj.marital_status,
                'educational_level': beneficiary_obj.educational_level,
                'health_status': beneficiary_obj.health_status,
                'disease_type': beneficiary_obj.disease_type,
                'work_status': beneficiary_obj.work_status,
                'employer': beneficiary_obj.employer,
                'death_date_father_husband': beneficiary_obj.death_date_father_husband,
                'washing_place': beneficiary_obj.washing_place,
                'is_benefiting': beneficiary_obj.is_benefiting,
                'received_at': beneficiary_obj.receivedAt,
                'reviewed_at': beneficiary_obj.reviewedAt,
                'bank_type': beneficiary_obj.bank_type,
                'bank_iban': beneficiary_obj.bank_iban,
                'family_issues': beneficiary_obj.family_issues,
                'family_needs': beneficiary_obj.family_needs,
                'dependent_list': dependent_data,
                'housing_info': housing_data,
                'income_expenses_info': income_expense_data,
                'attachments': beneficiary_attachment_list
            }
            return JsonResponse(data)
        except beneficiary.DoesNotExist:
            return JsonResponse({'error': 'Beneficiary not found'}, status=404)

    elif request.method == 'POST':
        pass


@login_required(login_url="/login")
def supporter_indiv(request):

    if request.method == 'GET':

        beneficiary_obj = beneficiary.objects.all().order_by('id')

        beneficiary_data = []

        in_ex_diff = 0

        for beneficiary_indiv in beneficiary_obj:
            # Retrieve beneficiary income and expenses information
            try:
                beneficiary_income_expenses_obj = beneficiary_income_expense.objects.filter(
                    beneficiary_id=beneficiary_indiv.id).first()

                if beneficiary_income_expenses_obj is not None:
                    in_ex_diff = beneficiary_income_expenses_obj.in_ex_diff
                else:
                    print("income is not available.")

            except ObjectDoesNotExist:
                beneficiary_income_expenses_obj = None

            # Collect the data and add them to the object
            beneficiary_data.append({
                'id': beneficiary_indiv.id,
                'gender': beneficiary_indiv.gender,
                'in_ex_diff': in_ex_diff,
                'category': beneficiary_indiv.category,
                'health_status': beneficiary_indiv.health_status,
                'age': beneficiary_indiv.age,
                'nationality': beneficiary_indiv.nationality,
            })

        # Prepare pagination
        # paginator = Paginator(beneficiary_data, IPP_SUPPORTER_FORM)
        # page_number = request.GET.get('page')
        # beneficiary_data = paginator.get_page(page_number)

        context = {
            'beneficiary_headers': ['#', 'الجنس', 'نسبة الاحتياج', 'التصنيف', 'الحالة الصحية', 'العمر', 'الجنسية'],
            'beneficiary_data': beneficiary_data,
        }
        return render(request, "main/supporter_form(indiv).html", context)

    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@csrf_exempt
@login_required(login_url="/login")
def supporter_indiv_post(request):
    if request.method == 'POST':

        # Retrieve form data
        # Get the data from the request
        post_data = request.POST
        files_data = request.FILES

        # Supporter information
        first_name = post_data.get('personalinfo_first_name', None)
        second_name = post_data.get('personalinfo_second_name', None)
        last_name = post_data.get('personalinfo_last_name', None)
        date_of_birth_data = post_data.get('personalinfo_date_of_birth', None)
        date_of_birth = None
        # Check if the date string exists and is not empty
        if date_of_birth_data:
            # Convert the date string to a date object
            date_of_birth = datetime.strptime(
                date_of_birth_data, '%Y-%m-%d').date()
        else:
            print("No valid date found in JSON")
        gender = post_data.get('personalinfo_gender', None)
        national_id = post_data.get('personalinfo_national_id', None)
        national_id_exp_date_data = post_data.get(
            'personalinfo_national_id_exp_date', None)
        national_id_exp_date = convert_to_date(
            national_id_exp_date_data)
        nationality = post_data.get('personalinfo_nationality', None)
        marital_status = post_data.get('personalinfo_marital_status', None)
        educational_level = post_data.get(
            'personalinfo_educational_level', None)
        work_status = post_data.get('personalinfo_work_status', None)
        employer = post_data.get('personalinfo_employer', None)
        phone_number = post_data.get('personalinfo_phone_number', None)
        email = post_data.get('personalinfo_email', None)

        # Supporter preferences
        was_sponsor = request.POST.get(
            'sponsorship_info_was_sponsor_check', None)
        status_notify = request.POST.get(
            'sponsorship_info_status_notify_check', None)
        invite_beneficiary = request.POST.get(
            'sponsorship_info_invite_beneficiary_check', None)
        visit_beneficiary = request.POST.get(
            'sponsorship_info_visit_beneficiary_check', None)

        supporter_obj = Supporter(
            first_name=first_name,
            second_name=second_name,
            last_name=last_name,
            date_of_birth=date_of_birth,
            gender=gender,
            national_id=national_id,
            national_id_exp_date=national_id_exp_date,
            nationality=nationality,
            marital_status=marital_status,
            educational_level=educational_level,
            work_status=work_status,
            employer=employer,
            phone_number=phone_number,
            email=email,
            status="غير مفعل",
            was_sponsor=was_sponsor,
            status_notify=status_notify,
            invite_beneficiary=invite_beneficiary,
            visit_beneficiary=visit_beneficiary,
        )
        supporter_obj.save()

        # Total price is the same either in personal or charity choice
        total_price = request.POST.get('total_price', None)

        # This field represent user option either to let the selection for the charity or do it by himself
        beneficiary_choice = request.POST.get('beneficiaries_choice')

        # Now you can access the data from form and selected rows
        if beneficiary_choice == "id_personal_choice":

            duration = request.POST.get('sponsorship_info_duration', None)
            donation_type = request.POST.get(
                'sponsorship_info_donation_type', None)

            # Retrieve selected rows' data
            try:
                selected_rows_data = json.loads(
                    request.POST.get('selectedRowsData'))
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON format in selectedRowsData'}, status=400)

            # A dictionary of beneficiary objects
            beneficiary_list = []

            for selected_row in selected_rows_data:

                beneficiary_list.append({
                    "id": selected_row[1],
                    "health_status": selected_row[5],
                    "category": selected_row[4],
                    "in_ex_diff": selected_row[3],
                    "gender": selected_row[2],
                    "age": selected_row[6],
                    "nationality": selected_row[7],
                })

            supporter_request_obj = Supporter_request(
                supporter=supporter_obj,
                status="انتظار",
                request_type="جديد",
                total_amount=total_price,
                selection_type="شخصي",
                duration=duration,
                donation_type=donation_type,
                beneficiary_list=beneficiary_list,
            )
            supporter_request_obj.save()

        else:

            orphan_number = request.POST.get(
                'charitychoice_orphan_number', None)
            orphan_donation_type = request.POST.get(
                'charitychoice_orphan_donation_type', None)
            widower_number = request.POST.get(
                'charitychoice_widower_number', None)
            widower_donation_type = request.POST.get(
                'charitychoice_widower_donation_type', None)

            supporter_request_obj = Supporter_request(
                supporter=supporter_obj,
                status="انتظار",
                request_type="جديد",
                total_amount=total_price,
                selection_type="الجمعية",
                orphan_number=orphan_number,
                orphan_donation_type=orphan_donation_type,
                widower_number=widower_number,
                widower_donation_type=widower_donation_type,
            )
            supporter_request_obj.save()

        bank_transfer_file = files_data.getlist(
            'file_bank_transfer')

        file_list = []

        # Loop for every file object to add to file_list
        for file_obj in bank_transfer_file:
            file_list.append(Supporter_request_attachment(
                supporter_request=supporter_request_obj,
                file_type="bank_transfer",
                file_object=file_obj,
            ))

        # Create the attachment objects for supporter request
        if file_list:
            Supporter_request_attachment.objects.bulk_create(file_list)

        # Return a JSON response as needed
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@group_required("Management")
@login_required(login_url="/login")
def supporter_request_details(request, supporter_id, s_request_id):

    supporter_obj = Supporter.objects.get(id=supporter_id)

    supporter_request_obj = Supporter_request.objects.filter(
        supporter=supporter_id, id=s_request_id).first()

    supporter_request_attachment_obj = Supporter_request_attachment.objects.filter(
        supporter_request=supporter_request_obj.id).all()

    # List of attachments for the supporter request
    supporter_request_attachment_list = []

    # Loop to add specific attributes to each attachment object
    for attachment in supporter_request_attachment_obj:
        # A variable that holds the attachment type in Arabic
        attachment_type_ar = ""

        if attachment.file_type == "bank_transfer":
            attachment_type_ar = "صورة إيصال الحوالة البنكية"
        else:
            attachment_type_ar = attachment.file_type

        supporter_request_attachment_list.append({
            'file_path': attachment.file_object.url,
            'file_extension': file_extension(attachment.file_object.url),
            'file_name': attachment.filename().split(".")[0],
            'file_size': attachment.file_size,
            'attachment_type': attachment_type_ar,
        })

    beneficiary_list = []

    # Retrieve information about each beneficiary for this request (in case of personal selection)
    for beneficiary_obj in supporter_request_obj.beneficiary_list:

        # Retrieve beneficiary information from DB
        beneficiary_temp = beneficiary.objects.get(
            id=beneficiary_obj['id'])

        # Retrieve beneficiary income and expenses information from DB
        beneficiary_income_expenses_temp = beneficiary_income_expense.objects.filter(
            beneficiary_id=beneficiary_temp.id).first()

        in_ex_diff = 0

        if beneficiary_income_expenses_temp is not None:
            in_ex_diff = beneficiary_income_expenses_temp.in_ex_diff

        beneficiary_list.append({
            "id": beneficiary_obj['id'],
            "full_name": (beneficiary_temp.first_name + ' ' + beneficiary_temp.second_name + ' ' + beneficiary_temp.last_name),
            "category": beneficiary_obj['category'],
            "in_ex_diff": in_ex_diff,
        })

    context = {
        "supporter": supporter_obj,
        "supporter_request": supporter_request_obj,
        "supporter_request_attachments": supporter_request_attachment_list,
        "beneficiary_list": beneficiary_list,
    }

    return render(request, "dashboard/supporter_request_details.html", context)


@group_required("Management")
@login_required(login_url="/login")
def supporter_request_confirm(request, supporter_id, s_request_id):
    if request.method == "POST":

        request_status = request.POST.get('request_status', None)
        request_comment = request.POST.get('request_comment', None)

        supporter_obj = Supporter.objects.get(id=supporter_id)

        supporter_request_obj = Supporter_request.objects.filter(
            supporter=supporter_id, id=s_request_id).first()

        # Update object data
        supporter_request_obj.status = request_status
        supporter_request_obj.comment = request_comment
        supporter_request_obj.reviewed_by = request.user
        supporter_request_obj.reviewed_at = datetime.now()

        # Save the changes
        supporter_request_obj.save()

        # Get the current date
        current_date = date.today()

        if supporter_request_obj.selection_type == 'شخصي' and (request_status == 'مقبول' or request_status == 'مكتمل'):

            # Get the date after the specified months
            date_after = current_date + \
                relativedelta(
                    months=+duration_factors[supporter_request_obj.duration])

            # Retrieve information about each beneficiary for this request (in case of personal selection)
            for beneficiary_obj in supporter_request_obj.beneficiary_list:

                # Retrieve beneficiary information from DB
                beneficiary_temp = beneficiary.objects.get(
                    id=beneficiary_obj['id'])

                # In case of other categories
                amount_donated_per_month = 400.0

                # In case of widower family, the total amount will change
                if beneficiary_temp.category == 'أسرة أرملة':
                    amount_donated_per_month = 600.0

                total_amount_per_beneficiary = amount_donated_per_month * \
                    duration_factors[supporter_request_obj.duration]

                # Create a relation between each beneficiary in the list of this supporter request
                sponsorship = Supporter_beneficiary_sponsorship(
                    amount_donated_monthly=amount_donated_per_month,
                    total_amount_donated=total_amount_per_beneficiary,
                    start_date=current_date,
                    end_date=date_after,
                    beneficiary=beneficiary_temp,
                    supporter=supporter_obj,
                )
                sponsorship.save()

            # Update the request status
            supporter_request_obj.status = "مكتمل"
            supporter_request_obj.save()

        messages.success(request, "لقد تم تحديث حالة طلب الداعم بنجاح!")
        return redirect("dashboard_supporters_requests")

    else:
        return JsonResponse({'status': 'error', 'message': 'Method Not Allowed'}, status=405)


@group_required("Management")
@login_required(login_url="/login")
def supporter_request_update(request, supporter_id, s_request_id):
    if request.method == "POST":

        request_status = request.POST.get('request_status', None)
        request_comment = request.POST.get('request_comment', None)

        supporter_obj = Supporter.objects.get(id=supporter_id)

        supporter_request_obj = Supporter_request.objects.filter(
            supporter=supporter_id, id=s_request_id).first()

        # Update object data
        supporter_request_obj.status = request_status
        supporter_request_obj.comment = request_comment
        supporter_request_obj.reviewed_by = request.user
        supporter_request_obj.reviewed_at = datetime.now()

        # Save the changes
        supporter_request_obj.save()

        messages.success(request, "لقد تم تحديث بيانات طلب الداعم بنجاح!")
        return redirect("dashboard_supporters_requests")


@login_required(login_url='/login')
def beneficiary_profile(request, user_id):
    # Get the logged-in user
    logged_in_user = request.user

    context = {}
    try:
        # Retrieve the user whose profile is being requested
        user_obj = CustomUser.objects.get(id=user_id)

        # Check if the logged-in user matches the requested user
        if logged_in_user != user_obj:
            messages.error(request, "ليس لديك الصلاحية اللازمة!")
            return redirect('home')
        
        # Get the groups associated with the user
        user_groups = user_obj.groups.first()

        # Convert date of birth to be populated in the template
        dob = date(
            user_obj.date_of_birth.year,
            user_obj.date_of_birth.month,
            user_obj.date_of_birth.day
        )

        # Convert national_id_exp_date to be populated in the template
        national_id_exp_date = ""
        if user_obj.national_id_exp_date is not None:
            national_id_exp_date = date(
                user_obj.national_id_exp_date.year,
                user_obj.national_id_exp_date.month,
                user_obj.national_id_exp_date.day
            )
            national_id_exp_date = int(time.mktime(national_id_exp_date.timetuple())) * 1000

        context = {
            'user_obj': user_obj,
            'user_dob': int(time.mktime(dob.timetuple())) * 1000,
            'user_national_id_exp_date': national_id_exp_date,
            'user_group': user_groups,
        }
    except ObjectDoesNotExist:
        messages.error(request, "المستخدم غير موجود!")
        return redirect('home')

    return render(request, 'main/beneficiary_profile.html', context)


@login_required(login_url='/login')
def beneficiary_profile_edit(request, user_id):
    
    if request.method == 'POST':

        data = request.POST

        first_name = data.get("first_name", None)
        second_name = data.get("second_name", None)
        last_name = data.get("last_name", None)
        username = data.get("username", None)
        date_of_birth = data.get("date_of_birth", None)
        national_id = data.get("national_id_edit", None)
        national_id_exp_date = data.get("national_id_exp_date_edit", None)
        
        # Check if the date string exists and is not empty
        if date_of_birth:
            # Convert the date string to a date object
            date_of_birth = datetime.strptime(
                date_of_birth, '%Y-%m-%d').date()
        else:
            print("No valid date found in JSON")
        gender = data.get("gender", None)
        nationality = data.get("nationality", None)

        user = get_object_or_404(CustomUser, id=user_id)

        user.first_name = first_name
        user.second_name = second_name
        user.last_name = last_name
        user.username = username
        user.date_of_birth = date_of_birth
        user.gender = gender
        user.nationality = nationality
        user.national_id = national_id
        user.national_id_exp_date = national_id_exp_date

        user.save()

        messages.success(request, "تم تعديل معلومات المستخدم بنجاح.")
        return redirect(reverse("beneficiary_profile", args=[user_id]))
    else:
        return render(request, "home.html")


@login_required(login_url='/login')
def beneficiary_requests(request, user_id):

    # Get the logged-in user
    logged_in_user = request.user

    context = {}
    try:
        # Retrieve the user whose profile is being requested
        user = CustomUser.objects.get(id=user_id)

        # Check if the logged-in user matches the requested user
        if logged_in_user != user:
            messages.error(request, "ليس لديك الصلاحية اللازمة!")
            return redirect('home')

        beneficiary_requests_list = Beneficiary_request.objects.filter(
            user=user.id).all()
        paginator = Paginator(beneficiary_requests_list,
                              5)
        page_number = request.GET.get('page')
        beneficiary_requests = paginator.get_page(page_number)
        context = {
            'user_info': user,
            'beneficiary_requests': beneficiary_requests,
            'request_count': len(beneficiary_requests)
        }
    except ObjectDoesNotExist:
        messages.error(request, "المستخدم غير موجود!")
        return redirect('home')

    return render(request, 'main/beneficiary_requests.html', context)


@login_required(login_url='/login')
def beneficiary_request_details(request, user_id):

    # Get the logged-in user
    logged_in_user = request.user

    context = {}
    try:
        # Retrieve the user whose profile is being requested
        user = CustomUser.objects.get(id=user_id)

        # Check if the logged-in user matches the requested user
        if logged_in_user != user:
            messages.error(request, "ليس لديك الصلاحية اللازمة!")
            return redirect('home')

        # Assuming `user_id` is the ID of the user you're checking for
        beneficiary_exists = beneficiary.objects.filter(user=user).exists()

        if not beneficiary_exists:
            messages.error(
                request, "ليس لديك ملف مستفيد لدينا!")
            return redirect('home')

        beneficiary_obj = beneficiary.objects.get(
            user=user)
        beneficiary_house_obj = beneficiary_house.objects.get(
            beneficiary_id=beneficiary_obj.id)
        beneficiary_income_expense_obj = beneficiary_income_expense.objects.get(
            beneficiary_id=beneficiary_obj.id)
        beneficiary_attachment_obj = Beneficiary_attachment.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        dependent_list = dependent.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        dependent_data = []

        for dependent_obj in dependent_list:

            # Initialize dependent income list with every dependent
            dependent_income_data = []

            # Retrieve the dependent income infomration
            dependent_income_list = Dependent_income.objects.filter(
                dependent=dependent_obj).all()

            # Add the data into the dependent income list
            for dependent_income_obj in dependent_income_list:
                dependent_income_data.append({
                    'id': dependent_income_obj.id,
                    'income_source': dependent_income_obj.source,
                    'income_amount': dependent_income_obj.amount,
                })

            dependent_data.append({
                'dependent_id': dependent_obj.id,
                'dependent_first_name': dependent_obj.first_name,
                'dependent_second_name': dependent_obj.second_name,
                'dependent_last_name': dependent_obj.last_name,
                'dependent_gender': dependent_obj.gender,
                'dependent_relationship': dependent_obj.relationship,
                'dependent_educational_status': dependent_obj.educational_status,
                'dependent_marital_status': dependent_obj.marital_status,
                'dependent_national_id': dependent_obj.national_id,
                'dependent_national_id_exp_date': dependent_obj.national_id_exp_date,
                'dependent_bank_iban': dependent_obj.bank_iban,
                'dependent_bank_type': dependent_obj.bank_type,
                'dependent_health_status': dependent_obj.health_status,
                'dependent_needs_type': dependent_obj.needs_type,
                'dependent_educational_degree': dependent_obj.educational_degree,
                'dependent_date_of_birth': dependent_obj.date_of_birth,
                'dependent_needs_description': dependent_obj.needs_description,
                'dependent_educational_level': dependent_obj.educational_level,
                'dependent_disease_type': dependent_obj.disease_type,
                'dependent_work_status': dependent_obj.work_status,
                'dependent_employer': dependent_obj.employer,
                'dependent_contribute_to_family_income': dependent_obj.contribute_to_family_income,
                'dependent_disability_check': dependent_obj.disability_check,
                'dependent_disability_type': dependent_obj.disability_type,
                'dependent_income_data': dependent_income_data,
            })

        beneficiary_attachment_list = []

        for attachment in beneficiary_attachment_obj:
            # A variable that holds the attachment type in Arabic
            attachment_type_ar = ""

            if attachment.file_type == "national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة"
            elif attachment.file_type == "national_address":
                attachment_type_ar = "العنوان الوطني"
            elif attachment.file_type == "dept_instrument":
                attachment_type_ar = "صك الدين"
            elif attachment.file_type == "pension_social_insurance":
                attachment_type_ar = "مشهد التقاعد أو التأمينات الاجتماعية"
            elif attachment.file_type == "father_husband_death_cert":
                attachment_type_ar = "شهادة الوفاة للزوج / الأب"
            elif attachment.file_type == "letter_from_prison":
                attachment_type_ar = "خطاب من السجن"
            elif attachment.file_type == "divorce_deed":
                attachment_type_ar = "صك الطلاق"
            elif attachment.file_type == "children_responsibility_deed":
                attachment_type_ar = "صك إعالة الأبناء"
            elif attachment.file_type == "other_files":
                attachment_type_ar = "مستندات أخرى"
            elif attachment.file_type == "lease_contract_title_deed":
                attachment_type_ar = "عقد الإيجار الالكتروني من منصة إيجار أو صك ملكية"
            elif attachment.file_type == "water_or_electricity_bills":
                attachment_type_ar = "الفواتير (كهرباء - ماء)"
            elif attachment.file_type == "dependent_national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة للمرافقين"
            elif attachment.file_type == "social_warranty_inquiry":
                attachment_type_ar = "مشهد الضمان الاجتماعي"
            else:
                attachment_type_ar = attachment.file_type

            beneficiary_attachment_list.append({
                'file_path': attachment.file_object.url,
                'file_extension': file_extension(attachment.file_object.url),
                'file_name': attachment.filename().split(".")[0],
                'file_size': attachment.file_size,
                'attachment_type': attachment_type_ar,
            })

        context = {
            'user_info': user,
            'beneficiary_requests': beneficiary_requests,
            'beneficiary': beneficiary_obj,
            'beneficiary_house': beneficiary_house_obj,
            'beneficiary_income_expense': beneficiary_income_expense_obj,
            'beneficiary_attachments': beneficiary_attachment_list,
            'dependent_list': dependent_data,
        }
    except ObjectDoesNotExist:
        messages.error(request, "المستخدم غير موجود!")
        return redirect('home')

    return render(request, "main/beneficiary_request_details.html", context)


@login_required(login_url="/login")
def beneficiary_request_update(request, user_id):

    # Get the logged-in user
    logged_in_user = request.user

    context = {}
    try:
        # Retrieve the user whose profile is being requested
        user = CustomUser.objects.get(id=user_id)

        # Check if the logged-in user matches the requested user
        if logged_in_user != user:
            messages.error(request, "ليس لديك الصلاحية اللازمة!")
            return redirect('home')

        # Assuming `user_id` is the ID of the user you're checking for
        beneficiary_exists = beneficiary.objects.filter(user=user).exists()

        if not beneficiary_exists:
            messages.error(
                request, "ليس لديك ملف مستفيد لدينا!")
            return redirect('home')
        # Get the last added beneficiary_request

        # # Check if there's existing beneficiary requests
        is_beneficiary_request = Beneficiary_request.objects.filter(
            user=user).exists()

        if is_beneficiary_request:
            last_beneficiary_request = Beneficiary_request.objects.filter(user=user).latest(
                'created_at')
            if last_beneficiary_request.status == "تحت المراجعة" or last_beneficiary_request.status == "انتظار":
                messages.error(
                    request, "لديك طلب سابق غير مكتمل! لا يمكنك إنشاء طلب جديد حتى يكتمل الطلب السابق.")
                return redirect('home')

        else:
            messages.error(
                request, "ليس لديك طلبات سابقة!!")
            return redirect('home')

        beneficiary_obj = beneficiary.objects.get(
            user=user_id)
        beneficiary_house_obj = beneficiary_house.objects.get(
            beneficiary_id=beneficiary_obj.id)
        beneficiary_income_expense_obj = beneficiary_income_expense.objects.get(
            beneficiary_id=beneficiary_obj.id)
        beneficiary_attachment_obj = Beneficiary_attachment.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        dependent_list = dependent.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        dependent_data = []

        for dependent_obj in dependent_list:

            # Initialize dependent income list with every dependent
            dependent_income_data = []

            # Retrieve the dependent income infomration
            dependent_income_list = Dependent_income.objects.filter(
                dependent=dependent_obj).all()

            # Add the data into the dependent income list
            for dependent_income_obj in dependent_income_list:
                dependent_income_data.append({
                    'id': dependent_income_obj.id,
                    'income_source': dependent_income_obj.source,
                    'income_amount': str(dependent_income_obj.amount),
                })

            dependent_data.append({
                'dependent_id': dependent_obj.id,
                'dependent_first_name': dependent_obj.first_name,
                'dependent_second_name': dependent_obj.second_name,
                'dependent_last_name': dependent_obj.last_name,
                'dependent_gender': dependent_obj.gender,
                'dependent_relationship': dependent_obj.relationship,
                'dependent_educational_status': dependent_obj.educational_status,
                'dependent_marital_status': dependent_obj.marital_status,
                'dependent_national_id': dependent_obj.national_id,
                'dependent_national_id_exp_date': dependent_obj.national_id_exp_date.strftime('%Y-%m-%d') if dependent_obj.national_id_exp_date else None,
                'dependent_health_status': dependent_obj.health_status,
                'dependent_needs_type': dependent_obj.needs_type,
                'dependent_educational_degree': dependent_obj.educational_degree,
                'dependent_date_of_birth': dependent_obj.date_of_birth.strftime('%Y-%m-%d'),
                'dependent_needs_description': dependent_obj.needs_description,
                'dependent_educational_level': dependent_obj.educational_level,
                'dependent_disease_type': dependent_obj.disease_type,
                'dependent_work_status': dependent_obj.work_status,
                'dependent_employer': dependent_obj.employer,
                'dependent_contribute_to_family_income': dependent_obj.contribute_to_family_income,
                'dependent_disability_check': dependent_obj.disability_check,
                'dependent_disability_type': dependent_obj.disability_type,
                'dependent_income_data': dependent_income_data,
            })

        beneficiary_attachment_list = []

        for attachment in beneficiary_attachment_obj:
            # A variable that holds the attachment type in Arabic
            attachment_type_ar = ""

            if attachment.file_type == "national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة"
            elif attachment.file_type == "national_address":
                attachment_type_ar = "العنوان الوطني"
            elif attachment.file_type == "dept_instrument":
                attachment_type_ar = "صك الدين"
            elif attachment.file_type == "pension_social_insurance":
                attachment_type_ar = "مشهد التقاعد أو التأمينات الاجتماعية"
            elif attachment.file_type == "father_husband_death_cert":
                attachment_type_ar = "شهادة الوفاة للزوج / الأب"
            elif attachment.file_type == "letter_from_prison":
                attachment_type_ar = "خطاب من السجن"
            elif attachment.file_type == "divorce_deed":
                attachment_type_ar = "صك الطلاق"
            elif attachment.file_type == "children_responsibility_deed":
                attachment_type_ar = "صك إعالة الأبناء"
            elif attachment.file_type == "other_files":
                attachment_type_ar = "مستندات أخرى"
            elif attachment.file_type == "lease_contract_title_deed":
                attachment_type_ar = "عقد الإيجار الالكتروني من منصة إيجار أو صك ملكية"
            elif attachment.file_type == "water_or_electricity_bills":
                attachment_type_ar = "الفواتير (كهرباء - ماء)"
            elif attachment.file_type == "dependent_national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة للمرافقين"
            elif attachment.file_type == "social_warranty_inquiry":
                attachment_type_ar = "مشهد الضمان الاجتماعي"
            else:
                attachment_type_ar = attachment.file_type

            beneficiary_attachment_list.append({
                'id': attachment.id,
                'file_path': attachment.file_object.url,
                'file_extension': file_extension(attachment.file_object.url),
                'file_name': attachment.filename().split(".")[0],
                'file_size': attachment.file_size,
                'attachment_type': attachment_type_ar,
            })

        # Convert date of birth to be populated in the template
        dob = date(
            user.date_of_birth.year,
            user.date_of_birth.month,
            user.date_of_birth.day
        )

        # Convert date of birth to be populated in the template
        national_id_exp_date = ""
        if user.national_id_exp_date is not None:
            national_id_exp_date = date(
                user.national_id_exp_date.year,
                user.national_id_exp_date.month,
                user.national_id_exp_date.day
            )
            national_id_exp_date = int(time.mktime(national_id_exp_date.timetuple())) * 1000
        

        death_date_father_husband = None
        if beneficiary_obj.death_date_father_husband is not None:
            death_date_father_husband = date(
                beneficiary_obj.death_date_father_husband.year,
                beneficiary_obj.death_date_father_husband.month,
                beneficiary_obj.death_date_father_husband.day
            )

        if death_date_father_husband is not None:
            context = {
                'user_info': user,
                'beneficiary_requests': beneficiary_requests,
                'beneficiary': beneficiary_obj,
                'beneficiary_dob': int(time.mktime(dob.timetuple())) * 1000,
                'beneficiary_death_date_father_husband': int(time.mktime(death_date_father_husband.timetuple())) * 1000,
                'beneficiary_national_id_exp_data': national_id_exp_date,
                'beneficiary_house': beneficiary_house_obj,
                'beneficiary_income_expense': beneficiary_income_expense_obj,
                'beneficiary_attachments': beneficiary_attachment_list,
                'dependent_list': json.dumps(dependent_data),
            }

        else:
            context = {
                'user_info': user,
                'beneficiary_requests': beneficiary_requests,
                'beneficiary': beneficiary_obj,
                'beneficiary_dob': int(time.mktime(dob.timetuple())) * 1000,
                'beneficiary_national_id_exp_data': national_id_exp_date,
                'beneficiary_house': beneficiary_house_obj,
                'beneficiary_income_expense': beneficiary_income_expense_obj,
                'beneficiary_attachments': beneficiary_attachment_list,
                'dependent_list': json.dumps(dependent_data),
            }

    except ObjectDoesNotExist:
        messages.error(request, "المستخدم غير موجود!")
        return redirect('home')

    return render(request, "main/beneficiary_request_update.html", context)


@csrf_exempt
@login_required(login_url="/login")
def beneficiary_request_update_confirm(request, user_id):

    # Get the logged-in user
    logged_in_user = request.user

    context = {}

    try:
        data = request.POST
        files = request.FILES

        # Retrieve the user whose profile is being requested
        user = CustomUser.objects.get(id=user_id)

        # Check if the logged-in user matches the requested user
        if logged_in_user != user:
            messages.error(request, "ليس لديك الصلاحية اللازمة!")
            return redirect('home')

        # Accessing the data for beneficiary
        category_personal = data.get('personalinfo_category', None)
        marital_status_personal = data.get('personalinfo_marital_status', None)
        educational_level_personal = data.get('personalinfo_educational_level', None)
        death_date_father_husband_personal = data.get(
            'personalinfo_death_date_father_husband', None)
        if death_date_father_husband_personal is not None:
            death_date_father_husband_personal = convert_to_date(
                death_date_father_husband_personal)
        washing_place_personal = data.get('personalinfo_washing_place', None)
        health_status_personal = data.get('personalinfo_health_status', None)
        disease_type_personal = data.get('personalinfo_disease_type', None)
        work_status_personal = data.get('personalinfo_work_status', None)
        employer_personal = data.get('personalinfo_employer', None)
        bank_type_personal = data.get('personalinfo_bank_type', None)
        bank_iban_personal = data.get('personalinfo_bank_iban', None)
        family_issues_personal = data.get('familyinfo_family_issues', None)
        family_needs_personal = data.get('familyinfo_needs_type', None)

        # Get beneficiary object using beneficiary_requests.beneficiary.id
        beneficiary_obj = beneficiary.objects.get(
            user=user)

        # Update object values
        beneficiary_obj.category = category_personal
        beneficiary_obj.marital_status = marital_status_personal
        beneficiary_obj.educational_level = educational_level_personal
        beneficiary_obj.death_date_father_husband = death_date_father_husband_personal
        beneficiary_obj.washing_place = washing_place_personal
        beneficiary_obj.health_status = health_status_personal
        beneficiary_obj.disease_type = disease_type_personal
        beneficiary_obj.work_status = work_status_personal
        beneficiary_obj.employer = employer_personal
        beneficiary_obj.bank_type = bank_type_personal
        beneficiary_obj.bank_iban = bank_iban_personal
        beneficiary_obj.family_issues = family_issues_personal
        beneficiary_obj.family_needs = family_needs_personal

        # Save object changes
        beneficiary_obj.save()

        # Accessing the data for beneficiary_house -----------------
        building_number = data.get('houseinfo_building_number', None)
        street_name = data.get('houseinfo_street_name', None)
        neighborhood = data.get('houseinfo_neighborhood', None)
        city = data.get('houseinfo_city', None)
        postal_code = data.get('houseinfo_postal_code', None)
        additional_number = data.get('houseinfo_additional_number', None)
        unit = data.get('houseinfo_unit', None)
        location_url = data.get('houseinfo_location_url', None)
        housing_type = data.get('houseinfo_housing_type', None)
        housing_ownership = data.get('houseinfo_housing_ownership', None)

        # Get corresponding beneficiary objects of other tables
        beneficiary_house_obj = beneficiary_house.objects.get(
            beneficiary_id=beneficiary_obj.id)

        # Update object values
        beneficiary_house_obj.building_number = building_number
        beneficiary_house_obj.street_name = street_name
        beneficiary_house_obj.neighborhood = neighborhood
        beneficiary_house_obj.city = city
        beneficiary_house_obj.postal_code = postal_code
        beneficiary_house_obj.additional_number = additional_number
        beneficiary_house_obj.unit = unit
        beneficiary_house_obj.location_url = location_url
        beneficiary_house_obj.housing_type = housing_type
        beneficiary_house_obj.housing_ownership = housing_ownership
        beneficiary_house_obj.beneficiary_id = beneficiary_obj

        # Save object changes
        beneficiary_house_obj.save()

        # Accessing the data for beneficiary_income_expense, and converting str into float type
        salary_in = float(data.get('incomeinfo_salary', None))
        social_insurance_in = float(
            data.get('incomeinfo_social_insurance', None))
        charity_in = float(data.get('incomeinfo_charity', None))
        social_warranty_in = float(
            data.get('incomeinfo_social_warranty', None))
        pension_agency_in = float(data.get('incomeinfo_pension_agency', None))
        citizen_account_in = float(
            data.get('incomeinfo_citizen_account', None))
        benefactor_in = float(data.get('incomeinfo_benefactor', None))
        other_in = float(data.get('incomeinfo_other', None))
        housing_rent_ex = float(data.get('expensesinfo_housing_rent', None))
        electricity_bills_ex = float(
            data.get('expensesinfo_electricity_bills', None))
        water_bills_ex = float(data.get('expensesinfo_water_bills', None))
        transportation_ex = float(
            data.get('expensesinfo_transportation', None))
        health_supplies_ex = float(
            data.get('expensesinfo_health_supplies', None))
        food_supplies_ex = float(data.get('expensesinfo_food_supplies', None))
        educational_supplies_ex = float(data.get(
            'expensesinfo_educational_supplies', None))
        proven_debts_ex = float(data.get('expensesinfo_proven_debts', None))
        other_ex = float(data.get('expensesinfo_other', None))

        beneficiary_income_expense_obj = beneficiary_income_expense.objects.get(
            beneficiary_id=beneficiary_obj.id)

        # Update object values
        beneficiary_income_expense_obj.salary_in = salary_in
        beneficiary_income_expense_obj.social_insurance_in = social_insurance_in
        beneficiary_income_expense_obj.charity_in = charity_in
        beneficiary_income_expense_obj.social_warranty_in = social_warranty_in
        beneficiary_income_expense_obj.pension_agency_in = pension_agency_in
        beneficiary_income_expense_obj.citizen_account_in = citizen_account_in
        beneficiary_income_expense_obj.benefactor_in = benefactor_in
        beneficiary_income_expense_obj.other_in = other_in
        beneficiary_income_expense_obj.housing_rent_ex = housing_rent_ex
        beneficiary_income_expense_obj.electricity_bills_ex = electricity_bills_ex
        beneficiary_income_expense_obj.water_bills_ex = water_bills_ex
        beneficiary_income_expense_obj.transportation_ex = transportation_ex
        beneficiary_income_expense_obj.health_supplies_ex = health_supplies_ex
        beneficiary_income_expense_obj.food_supplies_ex = food_supplies_ex
        beneficiary_income_expense_obj.educational_supplies_ex = educational_supplies_ex
        beneficiary_income_expense_obj.proven_debts_ex = proven_debts_ex
        beneficiary_income_expense_obj.other_ex = other_ex
        beneficiary_income_expense_obj.beneficiary_id = beneficiary_obj

        # Save object changes
        beneficiary_income_expense_obj.save()

        # Create a new beneficiary_request to have "under review" status
        new_beneficiary_request = Beneficiary_request(
            user=user,
            beneficiary=beneficiary_obj,
            status="انتظار",
            request_type="تحديث",
        )
        new_beneficiary_request.save()

        # beneficiary_attachment_obj = Beneficiary_attachment.objects.filter(
        #     beneficiary_id=beneficiary_obj.id).all()

        dependent_table = data.get('dependents-table', None)

        # print(data.get('dependents-table', None))
        # Parse the JSON string into a Python object
        try:
            dependents_list = json.loads(dependent_table)
        except json.JSONDecodeError:
            print("Error parsing JSON")
            dependents_list = []

        # Retrieve all dependents related to this beneficiary_obj
        beneficiary_dependents_db = dependent.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        if beneficiary_dependents_db.exists():
            # Iterate over the dependents
            for dependent_obj in beneficiary_dependents_db:
                # Check if the dependent exists in the response
                dependent_exists_in_response = any(
                    dependent_obj.national_id == dependent_data.get('nationalID') for dependent_data in dependents_list)

                # If the dependent doesn't exist in the response, delete it
                if not dependent_exists_in_response:
                    dependent_obj.delete()
        else:
            # Handle the case when there are no dependents for the beneficiary
            # For example, you can display a message or perform other actions
            print(
                "[Warning] - There are no dependents associated with this beneficiary.")

        for dep in dependents_list:

            # Extract the data for each field
            first_name = dep.get('firstName', '')
            second_name = dep.get('secondName', '')
            last_name = dep.get('lastName', '')
            gender = dep.get('gender', '')
            relationship = dep.get('relationship', '')
            educational_status = dep.get('educationalStatus', None)
            marital_status = dep.get('martialStatus', '')
            national_id = dep.get('nationalID', '')
            health_status = dep.get('healthStatus', None)
            needs_type = dep.get('needsType', '')
            educational_degree = dep.get('educationalDegree', '')

            date_of_birth = dep.get('dateOfBitrh', None)
            if date_of_birth is not None:
                date_of_birth = convert_to_date(date_of_birth)

            national_id_exp_date = dep.get(
                'nationalIDExpDate', None)
            if national_id_exp_date is not None:
                national_id_exp_date = convert_to_date(national_id_exp_date)

            needs_description = dep.get('needsDescription', '')
            educational_level = dep.get('educationalLevel', None)
            disease_type = dep.get('diseaseType', None)

            work_status = dep.get('workStatus', None)
            employer = dep.get('employer', None)
            contribute_to_family_income = dep.get(
                'contributeToFamilyIncome', None)
            disability_check = dep.get('disabilityCheck', None)
            disability_type = dep.get('disabilityType', None)

            dependent_income_table = json.loads(
                dep.get('dependentIncomeTable', []))

            # Check if national id exists before
            national_id_exist = dependent.objects.filter(
                national_id=national_id).exists()

            # In case of existing national id
            if national_id_exist:

                # Get the dependent object with existing national id
                dependent_obj = dependent.objects.filter(
                    national_id=national_id).first()

                # Update its information
                dependent_obj.first_name = first_name
                dependent_obj.second_name = second_name
                dependent_obj.last_name = last_name
                dependent_obj.gender = gender
                dependent_obj.relationship = relationship
                dependent_obj.date_of_birth = date_of_birth
                dependent_obj.national_id = national_id
                dependent_obj.national_id_exp_date = national_id_exp_date
                dependent_obj.marital_status = marital_status
                dependent_obj.educational_level = educational_level
                dependent_obj.educational_status = educational_status
                dependent_obj.health_status = health_status
                dependent_obj.disease_type = disease_type
                dependent_obj.needs_type = needs_type
                dependent_obj.educational_degree = educational_degree
                dependent_obj.needs_description = needs_description
                dependent_obj.work_status = work_status
                dependent_obj.employer = employer
                dependent_obj.contribute_to_family_income = contribute_to_family_income
                dependent_obj.disability_check = disability_check
                dependent_obj.disability_type = disability_type
                dependent_obj.beneficiary_id = beneficiary_obj

                # Save the changes
                dependent_obj.save()

                # Get list of all Dependent_income objects from the DB
                dependent_income_db_list = Dependent_income.objects.filter(
                    dependent=dependent_obj)

                # Delete existing dependent income records
                dependent_income_db_list.delete()

                # Store list of dependent income to create
                dependent_income_list = []

                # Traverse the list of dependent income from the request body
                for entry in dependent_income_table:
                    # Extract the monthly income and remove commas
                    income_amount_str = entry.get('income_amount', '')
                    income_amount = Decimal(income_amount_str.replace(',', ''))
                    income_source = entry.get('income_source', '')

                    # Initialize dependent income object
                    dependent_income_obj = Dependent_income(
                        source=income_source,
                        amount=income_amount,
                        dependent=dependent_obj
                    )
                    # Add the new dependent income object to the list
                    dependent_income_list.append(dependent_income_obj)

                # Save dependent income objects
                if dependent_income_list:
                    Dependent_income.objects.bulk_create(dependent_income_list)

            # In case of Non-existing national id
            else:

                # Create a new dependent object and save it to the database
                new_dependent = dependent(
                    first_name=first_name,
                    second_name=second_name,
                    last_name=last_name,
                    gender=gender,
                    relationship=relationship,
                    date_of_birth=date_of_birth,
                    national_id=national_id,
                    national_id_exp_date=national_id_exp_date,
                    marital_status=marital_status,
                    educational_level=educational_level,
                    educational_status=educational_status,
                    health_status=health_status,
                    disease_type=disease_type,
                    needs_type=needs_type,
                    educational_degree=educational_degree,
                    needs_description=needs_description,
                    work_status=work_status,
                    employer=employer,
                    contribute_to_family_income=contribute_to_family_income,
                    disability_check=disability_check,
                    disability_type=disability_type,
                    beneficiary_id=beneficiary_obj,
                )
                new_dependent.save()

                # Store list of income for dependent -------------------
                dependent_income_list = []

                for entry in dependent_income_table:
                    # Extract the monthly income and remove commas
                    income_amount_str = entry.get('income_amount', '')
                    income_amount = Decimal(
                        income_amount_str.replace(',', ''))
                    income_source = entry.get('income_source', '')

                    # Initialize dependent income list
                    dependent_income_obj = Dependent_income(
                        source=income_source,
                        amount=income_amount,
                        dependent=new_dependent
                    )
                    dependent_income_list.append(dependent_income_obj)

                # Save dependent income objects
                if dependent_income_list:
                    Dependent_income.objects.bulk_create(dependent_income_list)

        # Get filesToDelete array from request.POST (data)
        files_to_delete = json.loads(data.get('filesToDelete', '[]'))

        for file_id in files_to_delete:
            try:
                # Query Beneficiary_attachment object by its ID
                attachment = Beneficiary_attachment.objects.get(pk=file_id)

                # Delete file from storage
                if attachment.file_object:
                    # This will delete the file from storage
                    attachment.file_object.delete(save=False)

                # Delete the Beneficiary_attachment object
                attachment.delete()
            except Beneficiary_attachment.DoesNotExist:
                # Handle case where the Beneficiary_attachment object does not exist
                print('[Error] - Attchment to delete is not found!')

        # Get all the attachments of beneficiary
        national_id_file = files.get('fileBeneficiaryNationalID', None)
        national_address_file = files.get(
            'fileBeneficiaryNationalAddress', None)
        dept_instrument_file = files.getlist('fileDeptInstrument')
        pension_social_insurance_file = files.getlist(
            'filePensionOrSocialInsuranceInquiry')
        father_husband_death_certificate_file = files.get(
            'fileFatherOrHusbandDeathCertificate', None)
        letter_from_prison_file = files.getlist('fileLetterFromPrison')
        divorce_deed_file = files.get('fileDivorceDeed', None)
        children_responsibility_deed_file = files.getlist(
            'fileChildrenResponsibilityDeed')
        other_files = files.getlist('fileOther')
        lease_contract_or_title_deed_file = files.getlist(
            'fileLeaseContractOrTitleDeed')
        water_or_electricity_bills_file = files.getlist(
            'fileWaterOrElectricityBills')
        dependent_national_id_file = files.getlist(
            'fileNationalIDForBeneficiaryDependents')
        social_warranty_inquiry_file = files.getlist(
            'fileSocialWarrantyInquiry')

        print(files)

        # Store attachments of beneficiary -------------------
        # Store all file objects in a list
        file_list = []

        # Create beneficiary attachment for "national id"
        if national_id_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="national_id",
                file_object=national_id_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "national address"
        if national_address_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="national_address",
                file_object=national_address_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "dept instrument"
        for file_obj in dept_instrument_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="dept_instrument",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "pension or social insurance"
        for file_obj in pension_social_insurance_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="pension_social_insurance",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "father or husband death certificate"
        if father_husband_death_certificate_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="father_husband_death_cert",
                file_object=father_husband_death_certificate_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "letter from prison"
        for file_obj in letter_from_prison_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="letter_from_prison",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "Divorce Deed"
        if divorce_deed_file is not None:
            beneficiary_attachment_obj = Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="divorce_deed",
                file_object=divorce_deed_file,
            )
            file_list.append(beneficiary_attachment_obj)

        # Create beneficiary attachment for "children responsibility deed"
        for file_obj in children_responsibility_deed_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="children_responsibility_deed",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "other files"
        for file_obj in other_files:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="other_files",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "lease contract or title deed"
        for file_obj in lease_contract_or_title_deed_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="lease_contract_title_deed",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "water or electricity bills"
        for file_obj in water_or_electricity_bills_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="water_or_electricity_bills",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "dependent national id"
        for file_obj in dependent_national_id_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="dependent_national_id",
                file_object=file_obj
            ))

        # Create beneficiary attachment for "social warranty inquiry"
        for file_obj in social_warranty_inquiry_file:
            file_list.append(Beneficiary_attachment(
                beneficiary=beneficiary_obj,
                file_type="social_warranty_inquiry",
                file_object=file_obj
            ))

        # instead of creating and saving each file separately, store them in a list, and save them all at once.
        if file_list:
            Beneficiary_attachment.objects.bulk_create(file_list)

    except ObjectDoesNotExist:
        messages.error(request, "المستخدم غير موجود!")
        return redirect('home')

    return JsonResponse({'redirect': '/beneficiaries/requests/confirm_message/', 'user_id': user_id})
    # data = request.POST
    # print(data)
    # return JsonResponse({'redirect': '/confirmation', 'data': data})


@login_required(login_url="/login")
def confirm_beneficiary_request_update(request):

    return render(request, "main/beneficiary_update_request_confirm.html")


def validate_national_id_dependent(request, user_id):

    national_id = request.POST.get('national_id', None)

    if national_id is None:
        return HttpResponse("true")
    else:
        data = "false"

        d_data = not dependent.objects.filter(
            national_id=national_id).exists()
        s_data = not Supporter.objects.filter(
            national_id=national_id).exists()
        u_data = not CustomUser.objects.filter(
            national_id=national_id).exists()

        # in case of national_id doesn't exist before
        if d_data and s_data and u_data:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)


def validate_national_id_edit_user(request, user_id):

    # Check if the user_id is provided
    if user_id:
        user_obj = CustomUser.objects.filter(id=user_id).first()
    
        # Get base national id (registered id of the user in the DB)
        base_national_id = user_obj.national_id

        national_id = request.POST.get('national_id', None)

        if national_id is None:
            return HttpResponse("true")
        else:
            data = "false"

            d_data = not dependent.objects.filter(
                national_id=national_id).exists()
            s_data = not Supporter.objects.filter(
                national_id=national_id).exists()
            u_data = not CustomUser.objects.filter(
                national_id=national_id).exists()

            # in case of national_id doesn't exist before
            if d_data and s_data and u_data:
                data = "true"
            else:
                # in case of national_id exists before but it is equal to the base_national_id
                if base_national_id == national_id:
                    data = "true"
                else:
                    data = "false"

            return HttpResponse(data)
    else:
        return "User ID is not Found"
    

def validate_national_id_new_user(request):

    national_id = request.POST.get('national_id', None)

    if national_id is None:
        return HttpResponse("true")
    else:
        data = "false"

        d_data = not dependent.objects.filter(
            national_id=national_id).exists()
        s_data = not Supporter.objects.filter(
            national_id=national_id).exists()
        u_data = not CustomUser.objects.filter(
            national_id=national_id).exists()

        # in case of national_id doesn't exist before
        if d_data and s_data and u_data:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)


def validate_phonenumber_new_beneficiary(request, user_id):

    phonenumber = request.POST.get('phonenumber', None)

    if phonenumber is None:
        return HttpResponse("true")
    else:
        data = "false"

        s_data = not Supporter.objects.filter(
            phone_number=phonenumber).exists()
        u_data = not CustomUser.objects.filter(
            phonenumber=phonenumber).exists()

        # in case of national_id doesn't exist before
        if s_data and u_data:
            data = "true"
        else:
            data = "false"

        return HttpResponse(data)

# This handles that case of edit dependent where the national_id may equal to the current national_id.
# base_national_id is the national_id of the dependent before edit


def validate_national_id_edit_dependent(request, user_id):

    national_id = request.POST.get('national_id', None)
    base_nid = request.POST.get('base_nid', None)

    if national_id is None:
        return HttpResponse("true")
    else:
        data = "false"

        d_data = not dependent.objects.filter(
            national_id=national_id).exists()
        s_data = not Supporter.objects.filter(
            national_id=national_id).exists()
        u_data = not CustomUser.objects.filter(
            national_id=national_id).exists()

        # in case of national_id doesn't exist before
        if d_data and s_data and u_data:
            data = "true"
        else:
            # in case of national_id exists before but it is equal to the base_national_id
            if base_nid == national_id:
                data = "true"
            else:
                data = "false"

        return HttpResponse(data)


@group_required("Management")
@login_required(login_url="/login")
def supporter_beneficiary_sponsorship(request):

    context = {}

    sponsorships_list = Supporter_beneficiary_sponsorship.objects.all()

    context = {
        "sponsorships": sponsorships_list,
    }

    return render(request, "dashboard/sponsorships.html", context)


@group_required("Management")
@login_required(login_url="/login")
def add_sponsorship(request):

    if request.method == 'POST':

        supporter_request_id = request.POST.get('supporter_request', None)
        # Parse JSON data from the hidden input field
        beneficiary_data_json = request.POST.get('beneficiary_data')
        beneficiary_data = json.loads(beneficiary_data_json)

        supporter_request_obj = Supporter_request.objects.filter(
            pk=supporter_request_id).first()

        support_obj = Supporter.objects.filter(
            pk=supporter_request_obj.supporter.id).first()

        # Get the current date
        current_date = date.today()

        # Now you can access the beneficiary data as a list of dictionaries
        for beneficiary_obj in beneficiary_data:
            beneficiary_temp = beneficiary.objects.filter(
                pk=beneficiary_obj['id']).first()

            # In case of orphan family, the total amount will change
            if beneficiary_temp.category == 'أسرة أرملة':

                # Get the date after the specified months
                date_after = current_date + \
                    relativedelta(
                        months=+duration_factors[supporter_request_obj.widower_donation_type])

                amount_donated_per_month = 600.0

                total_amount_per_beneficiary = amount_donated_per_month * \
                    duration_factors[supporter_request_obj.widower_donation_type]

                sponsorship = Supporter_beneficiary_sponsorship(
                    total_amount_donated=supporter_request_obj.total_amount,
                    amount_donated_monthly=total_amount_per_beneficiary,
                    start_date=current_date,
                    end_date=date_after,
                    beneficiary=beneficiary_temp,
                    supporter=support_obj,
                )
                sponsorship.save()

            else:

                # Get the date after the specified months
                date_after = current_date + \
                    relativedelta(
                        months=+duration_factors[supporter_request_obj.orphan_donation_type])

                # In case of other categories
                amount_donated_per_month = 400.0

                total_amount_per_beneficiary = amount_donated_per_month * \
                    duration_factors[supporter_request_obj.orphan_donation_type]

                sponsorship = Supporter_beneficiary_sponsorship(
                    total_amount_donated=supporter_request_obj.total_amount,
                    amount_donated_monthly=total_amount_per_beneficiary,
                    start_date=current_date,
                    end_date=date_after,
                    beneficiary=beneficiary_temp,
                    supporter=support_obj,
                )
                sponsorship.save()

        # Update the request status
        supporter_request_obj.status = "مكتمل"
        supporter_request_obj.save()

        messages.success(request, "لقد تم إنشاء كفالة جديدة بنجاح!")
        return redirect("supporter_beneficiary_sponsorship")

    elif request.method == 'GET':

        # Retrieve only supporter requests with "charity selection"
        supporter_requests_list = Supporter_request.objects.filter(
            selection_type="الجمعية").all()

        beneficiary_list = beneficiary.objects.prefetch_related('user').all()

        context = {
            "supporter_requests": supporter_requests_list,
            "beneficiaries": beneficiary_list,
        }

        return render(request, "dashboard/add_sponsorship.html", context)

    return render(request, "dashboard/add_sponsorship.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_beneficiaries_list(request):

    context = {}

    beneficiaries_list = beneficiary.objects.all()

    context = {
        "beneficiaries": beneficiaries_list,
    }

    return render(request, "dashboard/beneficiaries.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_supporters_list(request):

    context = {}

    supporters_list = Supporter.objects.all()

    context = {
        "supporters": supporters_list,
    }

    return render(request, "dashboard/supporters.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_beneficiary_details(request, b_id):

    if request.method == "GET":

        beneficiary_obj = beneficiary.objects.get(id=b_id)

        user_obj = CustomUser.objects.get(id=beneficiary_obj.user_id)

        beneficiary_housing_obj = beneficiary_house.objects.filter(
            beneficiary_id=beneficiary_obj.id).first()

        housing_data = {}

        if beneficiary_housing_obj is not None:

            housing_data = {
                'building_number': beneficiary_housing_obj.building_number,
                'street_name': beneficiary_housing_obj.street_name,
                'neighborhood': beneficiary_housing_obj.neighborhood,
                'city': beneficiary_housing_obj.city,
                'postal_code': beneficiary_housing_obj.postal_code,
                'additional_number': beneficiary_housing_obj.additional_number,
                'unit': beneficiary_housing_obj.unit,
                'location_url': beneficiary_housing_obj.location_url,
                'housing_type': beneficiary_housing_obj.housing_type,
                'housing_ownership': beneficiary_housing_obj.housing_ownership
            }

        beneficiary_income_expense_obj = beneficiary_income_expense.objects.filter(
            beneficiary_id=beneficiary_obj.id).first()

        income_expense_data = {}

        if beneficiary_income_expense_obj is not None:
            income_expense_data = {
                'salary_in': beneficiary_income_expense_obj.salary_in,
                'social_insurance_in': beneficiary_income_expense_obj.social_insurance_in,
                'charity_in': beneficiary_income_expense_obj.charity_in,
                'social_warranty_in': beneficiary_income_expense_obj.social_warranty_in,
                'pension_agency_in': beneficiary_income_expense_obj.pension_agency_in,
                'citizen_account_in': beneficiary_income_expense_obj.citizen_account_in,
                'benefactor_in': beneficiary_income_expense_obj.benefactor_in,
                'other_in': beneficiary_income_expense_obj.other_in,
                'housing_rent_ex': beneficiary_income_expense_obj.housing_rent_ex,
                'electricity_bills_ex': beneficiary_income_expense_obj.electricity_bills_ex,
                'water_bills_ex': beneficiary_income_expense_obj.water_bills_ex,
                'transportation_ex': beneficiary_income_expense_obj.transportation_ex,
                'health_supplies_ex': beneficiary_income_expense_obj.health_supplies_ex,
                'food_supplies_ex': beneficiary_income_expense_obj.food_supplies_ex,
                'educational_supplies_ex': beneficiary_income_expense_obj.educational_supplies_ex,
                'proven_debts_ex': beneficiary_income_expense_obj.proven_debts_ex,
                'other_ex': beneficiary_income_expense_obj.other_ex
            }

        dependent_list = dependent.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        dependent_data = []

        for dependent_obj in dependent_list:

            # Initialize dependent income list with every dependent
            dependent_income_data = []

            # Retrieve the dependent income infomration
            dependent_income_list = Dependent_income.objects.filter(
                dependent=dependent_obj).all()

            # Add the data into the dependent income list
            for dependent_income_obj in dependent_income_list:
                dependent_income_data.append({
                    'income_source': dependent_income_obj.source,
                    'income_amount': dependent_income_obj.amount,
                })

            dependent_data.append({
                'dependent_id': dependent_obj.id,
                'dependent_first_name': dependent_obj.first_name,
                'dependent_second_name': dependent_obj.second_name,
                'dependent_last_name': dependent_obj.last_name,
                'dependent_gender': dependent_obj.gender,
                'dependent_relationship': dependent_obj.relationship,
                'dependent_educational_status': dependent_obj.educational_status,
                'dependent_marital_status': dependent_obj.marital_status,
                'dependent_national_id': dependent_obj.national_id,
                'dependent_national_id_exp_date': dependent_obj.national_id_exp_date,
                'dependent_bank_iban': dependent_obj.bank_iban,
                'dependent_bank_type': dependent_obj.bank_type,
                'dependent_health_status': dependent_obj.health_status,
                'dependent_needs_type': dependent_obj.needs_type,
                'dependent_educational_degree': dependent_obj.educational_degree,
                'dependent_date_of_birth': dependent_obj.date_of_birth,
                'dependent_needs_description': dependent_obj.needs_description,
                'dependent_educational_level': dependent_obj.educational_level,
                'dependent_disease_type': dependent_obj.disease_type,
                'dependent_work_status': dependent_obj.work_status,
                'dependent_employer': dependent_obj.employer,
                'dependent_contribute_to_family_income': dependent_obj.contribute_to_family_income,
                'dependent_disability_check': dependent_obj.disability_check,
                'dependent_disability_type': dependent_obj.disability_type,
                'dependent_income_data': dependent_income_data,
            })

        beneficiary_attachment_list = []

        attachments_list = Beneficiary_attachment.objects.filter(
            beneficiary_id=beneficiary_obj.id).all()

        for attachment in attachments_list:
            # A variable that holds the attachment type in Arabic
            attachment_type_ar = ""

            if attachment.file_type == "national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة"
            elif attachment.file_type == "national_address":
                attachment_type_ar = "العنوان الوطني"
            elif attachment.file_type == "dept_instrument":
                attachment_type_ar = "صك الدين"
            elif attachment.file_type == "pension_social_insurance":
                attachment_type_ar = "مشهد التقاعد أو التأمينات الاجتماعية"
            elif attachment.file_type == "father_husband_death_cert":
                attachment_type_ar = "شهادة الوفاة للزوج / الأب"
            elif attachment.file_type == "letter_from_prison":
                attachment_type_ar = "خطاب من السجن"
            elif attachment.file_type == "divorce_deed":
                attachment_type_ar = "صك الطلاق"
            elif attachment.file_type == "children_responsibility_deed":
                attachment_type_ar = "صك إعالة الأبناء"
            elif attachment.file_type == "other_files":
                attachment_type_ar = "مستندات أخرى"
            elif attachment.file_type == "lease_contract_title_deed":
                attachment_type_ar = "عقد الإيجار الالكتروني من منصة إيجار أو صك ملكية"
            elif attachment.file_type == "water_or_electricity_bills":
                attachment_type_ar = "الفواتير (كهرباء - ماء)"
            elif attachment.file_type == "dependent_national_id":
                attachment_type_ar = "صورة الهوية الوطنية/الإقامة للمرافقين"
            elif attachment.file_type == "social_warranty_inquiry":
                attachment_type_ar = "مشهد الضمان الاجتماعي"
            else:
                attachment_type_ar = attachment.file_type

            beneficiary_attachment_list.append({
                'file_path': attachment.file_object.url,
                'file_extension': file_extension(attachment.file_object.url),
                'file_name': attachment.filename().split(".")[0],
                'file_size': attachment.file_size,
                'attachment_type': attachment_type_ar,
            })
        # print("attachments: ", beneficiary_attachment_list)

        context = {
            'user_info': user_obj,
            'beneficiary': beneficiary_obj,
            'dependent_list': dependent_data,
            'beneficiary_house': housing_data,
            'beneficiary_income_expense': income_expense_data,
            'beneficiary_attachments': beneficiary_attachment_list
        }

        return render(request, "dashboard/beneficiary_details.html", context)

    else:
        pass


@group_required("Management")
@login_required(login_url="/login")
def dashboard_supporter_details(request, s_id):

    supporter = Supporter.objects.filter(id=s_id).first()

    print(supporter)

    context = {
        "supporter": supporter,
    }
    return render(request, "dashboard/supporter_details.html", context)


@group_required("Management")
@login_required(login_url="/login")
def dashboard_beneficiary_request_update(request, beneficiary_id, b_request_id):
    if request.method == "POST":

        request_status = request.POST.get('request_status', None)
        request_comment = request.POST.get('request_comment', None)

        beneficiary_request_obj = Beneficiary_request.objects.filter(
            beneficiary=beneficiary_id, id=b_request_id).first()

        # Update object data
        beneficiary_request_obj.status = request_status
        beneficiary_request_obj.comment = request_comment
        beneficiary_request_obj.reviewed_by = request.user
        beneficiary_request_obj.reviewed_at = datetime.now()

        # Save the changes
        beneficiary_request_obj.save()

        messages.success(request, "لقد تم تحديث حالة الطلب بنجاح!")
        return redirect("dashboard_beneficiaries_requests")

    else:
        return JsonResponse({'status': 'error', 'message': 'Method Not Allowed'}, status=405)


@group_required("Management")
@login_required(login_url='/login')
def dashboard_beneficiary_request_details(request, beneficiary_id, b_request_id):

    context = {}

    beneficiary_request_obj = Beneficiary_request.objects.filter(
        beneficiary=beneficiary_id, id=b_request_id).first()

    beneficiary_obj = beneficiary.objects.filter(
        pk=beneficiary_id).first()
    
    user_obj = CustomUser.objects.get(id=beneficiary_obj.user_id)

    beneficiary_house_obj = beneficiary_house.objects.get(
        beneficiary_id=beneficiary_obj.id)

    beneficiary_income_expense_obj = beneficiary_income_expense.objects.get(
        beneficiary_id=beneficiary_obj.id)

    beneficiary_attachment_obj = Beneficiary_attachment.objects.filter(
        beneficiary_id=beneficiary_obj.id).all()

    dependent_list = dependent.objects.filter(
        beneficiary_id=beneficiary_obj.id).all()

    dependent_data = []

    for dependent_obj in dependent_list:

        # Initialize dependent income list with every dependent
        dependent_income_data = []

        # Retrieve the dependent income infomration
        dependent_income_list = Dependent_income.objects.filter(
            dependent=dependent_obj).all()

        # Add the data into the dependent income list
        for dependent_income_obj in dependent_income_list:
            dependent_income_data.append({
                'id': dependent_income_obj.id,
                'income_source': dependent_income_obj.source,
                'income_amount': dependent_income_obj.amount,
            })

        dependent_data.append({
            'dependent_id': dependent_obj.id,
            'dependent_first_name': dependent_obj.first_name,
            'dependent_second_name': dependent_obj.second_name,
            'dependent_last_name': dependent_obj.last_name,
            'dependent_gender': dependent_obj.gender,
            'dependent_relationship': dependent_obj.relationship,
            'dependent_educational_status': dependent_obj.educational_status,
            'dependent_marital_status': dependent_obj.marital_status,
            'dependent_national_id': dependent_obj.national_id,
            'dependent_national_id_exp_date': dependent_obj.national_id_exp_date,
            'dependent_bank_iban': dependent_obj.bank_iban,
            'dependent_bank_type': dependent_obj.bank_type,
            'dependent_health_status': dependent_obj.health_status,
            'dependent_needs_type': dependent_obj.needs_type,
            'dependent_educational_degree': dependent_obj.educational_degree,
            'dependent_date_of_birth': dependent_obj.date_of_birth,
            'dependent_needs_description': dependent_obj.needs_description,
            'dependent_educational_level': dependent_obj.educational_level,
            'dependent_disease_type': dependent_obj.disease_type,
            'dependent_work_status': dependent_obj.work_status,
            'dependent_employer': dependent_obj.employer,
            'dependent_contribute_to_family_income': dependent_obj.contribute_to_family_income,
            'dependent_disability_check': dependent_obj.disability_check,
            'dependent_disability_type': dependent_obj.disability_type,
            'dependent_income_data': dependent_income_data,
        })

    beneficiary_attachment_list = []

    for attachment in beneficiary_attachment_obj:
        # A variable that holds the attachment type in Arabic
        attachment_type_ar = ""

        if attachment.file_type == "national_id":
            attachment_type_ar = "صورة الهوية الوطنية/الإقامة"
        elif attachment.file_type == "national_address":
            attachment_type_ar = "العنوان الوطني"
        elif attachment.file_type == "dept_instrument":
            attachment_type_ar = "صك الدين"
        elif attachment.file_type == "pension_social_insurance":
            attachment_type_ar = "مشهد التقاعد أو التأمينات الاجتماعية"
        elif attachment.file_type == "father_husband_death_cert":
            attachment_type_ar = "شهادة الوفاة للزوج / الأب"
        elif attachment.file_type == "letter_from_prison":
            attachment_type_ar = "خطاب من السجن"
        elif attachment.file_type == "divorce_deed":
            attachment_type_ar = "صك الطلاق"
        elif attachment.file_type == "children_responsibility_deed":
            attachment_type_ar = "صك إعالة الأبناء"
        elif attachment.file_type == "other_files":
            attachment_type_ar = "مستندات أخرى"
        elif attachment.file_type == "lease_contract_title_deed":
            attachment_type_ar = "عقد الإيجار الالكتروني من منصة إيجار أو صك ملكية"
        elif attachment.file_type == "water_or_electricity_bills":
            attachment_type_ar = "الفواتير (كهرباء - ماء)"
        elif attachment.file_type == "dependent_national_id":
            attachment_type_ar = "صورة الهوية الوطنية/الإقامة للمرافقين"
        elif attachment.file_type == "social_warranty_inquiry":
            attachment_type_ar = "مشهد الضمان الاجتماعي"
        else:
            attachment_type_ar = attachment.file_type

        beneficiary_attachment_list.append({
            'file_path': attachment.file_object.url,
            'file_extension': file_extension(attachment.file_object.url),
            'file_name': attachment.filename().split(".")[0],
            'file_size': attachment.file_size,
            'attachment_type': attachment_type_ar,
        })

    context = {
        'user_info': user_obj,
        'beneficiary': beneficiary_obj,
        'beneficiary_request': beneficiary_request_obj,
        'beneficiary_house': beneficiary_house_obj,
        'beneficiary_income_expense': beneficiary_income_expense_obj,
        'beneficiary_attachments': beneficiary_attachment_list,
        'dependent_list': dependent_data,
    }

    return render(request, "dashboard/beneficiary_request_details.html", context)


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_users(request):

    context = {}

    #  used to fetch all related groups for each user efficiently. This way, when you iterate over the users_list in your template, you can access the groups associated with each user.
    users_list = CustomUser.objects.prefetch_related('groups').all()

    context = {
        'users_list': users_list,
    }

    return render(request, "dashboard/users_list.html", context)


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_profile(request, user_id):

    context = {}

    user_obj = CustomUser.objects.filter(pk=user_id).first()

    if user_obj:
        # Get the groups associated with the user
        user_groups = user_obj.groups.first()

        # Convert date of birth to be populated in the template
        dob = date(
            user_obj.date_of_birth.year,
            user_obj.date_of_birth.month,
            user_obj.date_of_birth.day
        )

        # Convert national_id_exp_date to be populated in the template
        national_id_exp_date = ""
        if user_obj.national_id_exp_date is not None:
            national_id_exp_date = date(
                user_obj.national_id_exp_date.year,
                user_obj.national_id_exp_date.month,
                user_obj.national_id_exp_date.day
            )
            national_id_exp_date = int(time.mktime(national_id_exp_date.timetuple())) * 1000

        context = {
            'user_obj': user_obj,
            'user_dob': int(time.mktime(dob.timetuple())) * 1000,
            'user_national_id_exp_date': national_id_exp_date,
            'user_group': user_groups,
        }
    else:
        # Handle case when user is not found
        context['error_message'] = "User not found."

    return render(request, "dashboard/user_profile.html", context)


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_delete(request, user_id):

    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        user.delete()
        messages.success(request, "تم حذف المستخدم بنجاح.")
        return redirect("dashboard_users")
    else:
        return render(request, "dashboard/users_list.html")


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_edit_basic_info(request, user_id):

    if request.method == 'POST':

        data = request.POST

        first_name = data.get("first_name", None)
        second_name = data.get("second_name", None)
        last_name = data.get("last_name", None)
        username = data.get("username", None)
        date_of_birth = data.get("date_of_birth", None)
        # Check if the date string exists and is not empty
        if date_of_birth:
            # Convert the date string to a date object
            date_of_birth = datetime.strptime(
                date_of_birth, '%Y-%m-%d').date()
        else:
            print("No valid date found in JSON")
        gender = data.get("gender", None)
        nationality = data.get("nationality", None)
        national_id = data.get("national_id", None)
        national_id_exp_date = data.get("national_id_exp_date", None)
        if national_id_exp_date:
            # Convert the date string to a date object
            national_id_exp_date = datetime.strptime(
                national_id_exp_date, '%Y-%m-%d').date()

        user = get_object_or_404(CustomUser, id=user_id)

        user.first_name = first_name
        user.second_name = second_name
        user.last_name = last_name
        user.username = username
        user.date_of_birth = date_of_birth
        user.gender = gender
        user.nationality = nationality
        user.national_id = national_id
        user.national_id_exp_date = national_id_exp_date

        user.save()

        messages.success(request, "تم تعديل معلومات المستخدم بنجاح.")
        return redirect(reverse("dashboard_user_profile", args=[user_id]))
    else:
        return render(request, "dashboard/users_list.html")
    

@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_edit_email(request, user_id):

    if request.method == 'POST':

        data = request.POST

        email = data.get("update_email", None)

        user = get_object_or_404(CustomUser, id=user_id)

        user.email = email

        user.save()

        messages.success(request, "تم تعديل معلومات المستخدم بنجاح.")
        return redirect(reverse("dashboard_user_profile", args=[user_id]))
    else:
        return render(request, "dashboard/users_list.html")


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_validate_username(request):

    username = request.POST.get('username', None)
    b_username = request.POST.get('base_username', None)

    if username is None:
        return HttpResponse("true")
    else:
        data = "false"

        cu_data = not CustomUser.objects.filter(
            username=username).exists()

        if cu_data:
            data = "true"
        else:
            # in case of username exists before but it is equal to the base_username
            if username == b_username:
                data = "true"
            else:
                data = "false"

        return HttpResponse(data)
    

@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_validate_national_id(request):

    national_id = request.POST.get('national_id', None)
    b_national_id = request.POST.get('base_national_id', None)

    if national_id is None:
        return HttpResponse("true")
    else:
        data = "false"

        cu_data = not CustomUser.objects.filter(
            national_id=national_id).exists()

        if cu_data:
            data = "true"
        else:
            # in case of national_id exists before but it is equal to the base_national_id
            if national_id == b_national_id:
                data = "true"
            else:
                data = "false"

        return HttpResponse(data)
    

@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_validate_email(request):

    email = request.POST.get('email', None)
    b_email= request.POST.get('base_email', None)

    if email is None:
        return HttpResponse("true")
    else:
        data = "false"

        cu_data = not CustomUser.objects.filter(
            email=email).exists()

        if cu_data:
            data = "true"
        else:
            # in case of email exists before but it is equal to the base_email
            if email == b_email:
                data = "true"
            else:
                data = "false"

        return HttpResponse(data)


@group_required("Admin")
@login_required(login_url='/login')
def dashboard_user_edit_role(request, user_id):

    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':

        data = request.POST
        new_role_name = data.get("new_role", None)

        # Get the group corresponding to the selected role
        new_role_group = Group.objects.filter(
            name__iexact=new_role_name).first()

        # Remove the user from all previous groups
        user.groups.clear()

        # Add the user to the new role group
        user.groups.add(new_role_group)

        messages.success(request, "تم تعديل دور المستخدم بنجاح.")
        return redirect(reverse("dashboard_user_profile", args=[user_id]))
    else:
        messages.error(request, "لقد حدث خطأ غير متوقع.")
        return redirect("dashboard_users")


@group_required("Management")
@login_required(login_url='/login')
def dashboard_support_operations(request):

    if request.method == 'GET':

        support_operations_list = Support_operation.objects.prefetch_related(
            'beneficiary').all()

        beneficiaries_list = beneficiary.objects.all()

        paginator = Paginator(support_operations_list, IPP_DASHBOARD_REQUESTS)
        page_number = request.GET.get('page')
        support_operations_list = paginator.get_page(page_number)

        context = {
            "support_operations": support_operations_list,
            "beneficiaries": beneficiaries_list,
        }

        return render(request, "dashboard/support_operations.html", context)
    else:
        messages.error(request, "لقد حدث خطأ غير متوقع.")
        return redirect("dashboard")


@group_required("Management")
@login_required(login_url='/login')
def dashboard_add_support_operation(request):

    if request.method == 'POST':

        data = request.POST
        files = request.FILES

        support_operation_type = data.get("support_operation_type", None)
        beneficiary_id = data.get("beneficiary", None)
        total_amount = data.get("total_amount", None)
        notes = data.get("notes", None)

        beneficiary_obj = beneficiary.objects.filter(pk=beneficiary_id).first()

        support_operation_obj = Support_operation(
            beneficiary=beneficiary_obj,
            support_type=support_operation_type,
            notes=notes,
            total_amount=total_amount,
        )
        support_operation_obj.save()

        support_operation_attachments = files.getlist(
            'support_operation_attachment')

        file_list = []

        # Loop for every file object to add to file_list
        for file_obj in support_operation_attachments:
            file_list.append(Support_operation_attachment(
                support_operation=support_operation_obj,
                file_type="general",
                file_object=file_obj,
            ))

        # Create the attachment objects for supporter operations
        if file_list:
            Support_operation_attachment.objects.bulk_create(file_list)

        messages.success(request, "تم إضافة عملية الدعم بنجاح.")
        return redirect("dashboard_support_operations")
    else:
        messages.error(request, "لقد حدث خطأ غير متوقع.")
        return redirect("dashboard")


@group_required("Management")
@login_required(login_url='/login')
def dashboard_field_visits(request):

    field_visits_list = Field_visit.objects.all()

    beneficiaries_list = beneficiary.objects.all()

    paginator = Paginator(field_visits_list, IPP_DASHBOARD_REQUESTS)
    page_number = request.GET.get('page')
    field_visits_list = paginator.get_page(page_number)

    context = {
        'field_visits_list': field_visits_list,
        "beneficiaries": beneficiaries_list,
    }

    return render(request, "dashboard/field_visit.html", context)


@group_required("Management")
@login_required(login_url='/login')
def dashboard_add_field_visit(request):

    if request.method == 'POST':

        data = request.POST
        files = request.FILES

        beneficiary_id = data.get("beneficiary", None)
        report_after_visit = data.get("report_after_visit", None)

        beneficiary_obj = beneficiary.objects.filter(pk=beneficiary_id).first()

        field_visit_obj = Field_visit(
            beneficiary=beneficiary_obj,
            specialist=request.user,
            visit_type="",
            report_after_visit=report_after_visit,
        )
        field_visit_obj.save()

        field_visit_attachments = files.getlist(
            'field_visit_attachment')

        file_list = []

        # Loop for every file object to add to file_list
        for file_obj in field_visit_attachments:
            file_list.append(Field_visit_attachment(
                field_visit=field_visit_obj,
                file_type="general",
                file_object=file_obj,
            ))

        # Create the attachment objects for Field_visit_attachment
        if file_list:
            Field_visit_attachment.objects.bulk_create(file_list)

        messages.success(request, "تم إضافة زيارة ميدانية بنجاح.")
        return redirect("dashboard_field_visits")
    else:
        messages.error(request, "لقد حدث خطأ غير متوقع.")
        return redirect("dashboard")